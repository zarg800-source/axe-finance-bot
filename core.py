"""
core.py — Shared business logic for Axe Finance
=================================================
This module holds everything the finance app needs that used to live inside
main.py's Telegram bot: category/account keyword maps, category & account
detection, database init/access, recurring-subscription auto-processing,
and the monthly Google Drive Excel backup.

It has ZERO Telegram dependencies — Axe Finance is now a pure web app
(Flask + the dashboard + the Axe chat/receipt-scanning assistant).
"""

import os
import re
import io
import json
import sqlite3
import logging
import base64
import requests
from datetime import datetime, timedelta, date
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

logger = logging.getLogger(__name__)

# ─── Config ─────────────────────────────────────────────────────────────────
AUTHORIZED_USER_ID = int(os.environ.get('AUTHORIZED_USER_ID', '0'))
DATABASE_NAME = os.path.join('/data', 'finance.db')
BANGKOK_TZ = pytz.timezone('Asia/Bangkok')

# ─── Google Drive backup (optional) ──────────────────────────────────────────
try:
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
    from google.oauth2 import service_account
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False
    logger.warning("Google API packages not installed. Drive backup disabled.")


CATEGORIES = {
    # ─── 🍜 Food & Drinks ─────────────────────────────────────────────────
    'food': ('Food & Drinks', '🍜'),
    'food & drinks': ('Food & Drinks', '🍜'),
    'restaurant': ('Food & Drinks', '🍜'),
    'street food': ('Food & Drinks', '🍜'),
    'lunch': ('Food & Drinks', '🍜'),
    'dinner': ('Food & Drinks', '🍜'),
    'breakfast': ('Food & Drinks', '🍜'),
    'brunch': ('Food & Drinks', '🍜'),
    'snack': ('Food & Drinks', '🍜'),
    'supper': ('Food & Drinks', '🍜'),
    # Fast food & chains
    'kfc': ('Food & Drinks', '🍜'),
    'bonchon': ('Food & Drinks', '🍜'),
    'mcdonalds': ('Food & Drinks', '🍜'),
    'mcdonald': ('Food & Drinks', '🍜'),
    'mcd': ('Food & Drinks', '🍜'),
    'burger king': ('Food & Drinks', '🍜'),
    'pizza hut': ('Food & Drinks', '🍜'),
    'pizza company': ('Food & Drinks', '🍜'),
    'the pizza company': ('Food & Drinks', '🍜'),
    'dominos': ('Food & Drinks', '🍜'),
    'subway': ('Food & Drinks', '🍜'),
    'mos burger': ('Food & Drinks', '🍜'),
    'texas chicken': ('Food & Drinks', '🍜'),
    'popeyes': ('Food & Drinks', '🍜'),
    'wendys': ('Food & Drinks', '🍜'),
    'five guys': ('Food & Drinks', '🍜'),
    'shake shack': ('Food & Drinks', '🍜'),
    # Thai chains & restaurants
    'mk': ('Food & Drinks', '🍜'),
    'mk suki': ('Food & Drinks', '🍜'),
    'mk restaurant': ('Food & Drinks', '🍜'),
    'sizzler': ('Food & Drinks', '🍜'),
    's&p': ('Food & Drinks', '🍜'),
    'bar b q plaza': ('Food & Drinks', '🍜'),
    'bar-b-q plaza': ('Food & Drinks', '🍜'),
    'bbq plaza': ('Food & Drinks', '🍜'),
    'shabushi': ('Food & Drinks', '🍜'),
    'fuji': ('Food & Drinks', '🍜'),
    'fuji restaurant': ('Food & Drinks', '🍜'),
    'yayoi': ('Food & Drinks', '🍜'),
    'oishi': ('Food & Drinks', '🍜'),
    'oishi ramen': ('Food & Drinks', '🍜'),
    'oishi grand': ('Food & Drinks', '🍜'),
    'coco ichibanya': ('Food & Drinks', '🍜'),
    'pepper lunch': ('Food & Drinks', '🍜'),
    'hachiban ramen': ('Food & Drinks', '🍜'),
    'gyukatsu': ('Food & Drinks', '🍜'),
    'sukishi': ('Food & Drinks', '🍜'),
    'after you': ('Food & Drinks', '🍜'),
    'swensens': ('Food & Drinks', '🍜'),
    'dairy queen': ('Food & Drinks', '🍜'),
    'dq': ('Food & Drinks', '🍜'),
    'baskin robbins': ('Food & Drinks', '🍜'),
    'cold stone': ('Food & Drinks', '🍜'),
    'krispy kreme': ('Food & Drinks', '🍜'),
    'mister donut': ('Food & Drinks', '🍜'),
    'dunkin donuts': ('Food & Drinks', '🍜'),
    'dunkin': ('Food & Drinks', '🍜'),
    'auntie annes': ('Food & Drinks', '🍜'),
    'donut': ('Food & Drinks', '🍜'),
    'donuts': ('Food & Drinks', '🍜'),
    'golden donuts': ('Food & Drinks', '🍜'),
    'mr donut': ('Food & Drinks', '🍜'),
    'bread': ('Food & Drinks', '🍜'),
    'toast': ('Food & Drinks', '🍜'),
    'bakery': ('Food & Drinks', '🍜'),
    'cake': ('Food & Drinks', '🍜'),
    'pastry': ('Food & Drinks', '🍜'),
    'croissant': ('Food & Drinks', '🍜'),
    'waffle': ('Food & Drinks', '🍜'),
    'pancake': ('Food & Drinks', '🍜'),
    'dessert': ('Food & Drinks', '🍜'),
    'sweets': ('Food & Drinks', '🍜'),
    'candy': ('Food & Drinks', '🍜'),
    'chocolate': ('Food & Drinks', '🍜'),
    'ice cream': ('Food & Drinks', '🍜'),
    'big mac': ('Food & Drinks', '🍜'),
    'whopper': ('Food & Drinks', '🍜'),
    'fried chicken': ('Food & Drinks', '🍜'),
    'wings': ('Food & Drinks', '🍜'),
    'kebab': ('Food & Drinks', '🍜'),
    'shawarma': ('Food & Drinks', '🍜'),
    'crepe': ('Food & Drinks', '🍜'),
    'taco': ('Food & Drinks', '🍜'),
    'burrito': ('Food & Drinks', '🍜'),
    'poke': ('Food & Drinks', '🍜'),
    'acai': ('Food & Drinks', '🍜'),
    'kanom': ('Food & Drinks', '🍜'),
    'khanom': ('Food & Drinks', '🍜'),
    'cha yen': ('Food & Drinks', '🍜'),
    'mango': ('Food & Drinks', '🍜'),
    'durian': ('Food & Drinks', '🍜'),
    'fruit': ('Food & Drinks', '🍜'),
    'thirteen coins': ('Food & Drinks', '🍜'),
    'the terrace': ('Food & Drinks', '🍜'),
    'sushi': ('Food & Drinks', '🍜'),
    'ramen': ('Food & Drinks', '🍜'),
    'som tam': ('Food & Drinks', '🍜'),
    'somtam': ('Food & Drinks', '🍜'),
    'pad thai': ('Food & Drinks', '🍜'),
    'noodle': ('Food & Drinks', '🍜'),
    'noodles': ('Food & Drinks', '🍜'),
    'rice': ('Food & Drinks', '🍜'),
    'chicken': ('Food & Drinks', '🍜'),
    'pork': ('Food & Drinks', '🍜'),
    'beef': ('Food & Drinks', '🍜'),
    'fish': ('Food & Drinks', '🍜'),
    'seafood': ('Food & Drinks', '🍜'),
    'pizza': ('Food & Drinks', '🍜'),
    'burger': ('Food & Drinks', '🍜'),
    'hotdog': ('Food & Drinks', '🍜'),
    'sandwich': ('Food & Drinks', '🍜'),
    'salad': ('Food & Drinks', '🍜'),
    'steak': ('Food & Drinks', '🍜'),
    'buffet': ('Food & Drinks', '🍜'),
    'hotpot': ('Food & Drinks', '🍜'),
    'shabu': ('Food & Drinks', '🍜'),
    'yakiniku': ('Food & Drinks', '🍜'),
    'bbq': ('Food & Drinks', '🍜'),
    'grill': ('Food & Drinks', '🍜'),
    'dim sum': ('Food & Drinks', '🍜'),
    'dimsum': ('Food & Drinks', '🍜'),
    'congee': ('Food & Drinks', '🍜'),
    'jok': ('Food & Drinks', '🍜'),
    'khao man gai': ('Food & Drinks', '🍜'),
    'boat noodle': ('Food & Drinks', '🍜'),
    'thai food': ('Food & Drinks', '🍜'),
    'japanese food': ('Food & Drinks', '🍜'),
    'korean food': ('Food & Drinks', '🍜'),
    'chinese food': ('Food & Drinks', '🍜'),
    'indian food': ('Food & Drinks', '🍜'),
    'italian food': ('Food & Drinks', '🍜'),
    'western food': ('Food & Drinks', '🍜'),
    # Drinks
    'pepsi': ('Food & Drinks', '🍜'),
    'coke': ('Food & Drinks', '🍜'),
    'coca cola': ('Food & Drinks', '🍜'),
    'sprite': ('Food & Drinks', '🍜'),
    'fanta': ('Food & Drinks', '🍜'),
    'est': ('Food & Drinks', '🍜'),
    'oishi tea': ('Food & Drinks', '🍜'),
    'ichitan': ('Food & Drinks', '🍜'),
    'bubble tea': ('Food & Drinks', '🍜'),
    'boba': ('Food & Drinks', '🍜'),
    'milk tea': ('Food & Drinks', '🍜'),
    'cha tra mue': ('Food & Drinks', '🍜'),
    'thai tea': ('Food & Drinks', '🍜'),
    'smoothie': ('Food & Drinks', '🍜'),
    'juice': ('Food & Drinks', '🍜'),
    # Delivery
    'grab food': ('Food & Drinks', '🍜'),
    'grabfood': ('Food & Drinks', '🍜'),
    'food panda': ('Food & Drinks', '🍜'),
    'foodpanda': ('Food & Drinks', '🍜'),
    'lineman': ('Food & Drinks', '🍜'),
    'line man': ('Food & Drinks', '🍜'),
    'robinhood': ('Food & Drinks', '🍜'),
    # QSR (Quick Service Restaurant) companies on receipts
    'qsr of asia': ('Food & Drinks', '🍜'),
    'qsr': ('Food & Drinks', '🍜'),
    'minor food': ('Food & Drinks', '🍜'),
    'the minor food': ('Food & Drinks', '🍜'),
    'central restaurants': ('Food & Drinks', '🍜'),
    'crg': ('Food & Drinks', '🍜'),
    # Thai company names that appear on bank receipts ("To" field)
    'yum restaurants': ('Food & Drinks', '🍜'),
    'yum': ('Food & Drinks', '🍜'),
    'mcdonald': ('Food & Drinks', '🍜'),
    'mcdonalds': ('Food & Drinks', '🍜'),
    'cpall': ('Food & Drinks', '🍜'),
    'cp all': ('Food & Drinks', '🍜'),
    'central food retail': ('Groceries', '🛒'),
    'central food': ('Groceries', '🛒'),
    'big c supercenter': ('Groceries', '🛒'),
    'siam makro': ('Groceries', '🛒'),
    'ek-chai distribution': ('Groceries', '🛒'),
    'cp axtra': ('Groceries', '🛒'),
    'true digital group': ('Subscriptions', '📱'),
    'true digital': ('Subscriptions', '📱'),
    'true corp': ('Subscriptions', '📱'),
    'true corporation': ('Subscriptions', '📱'),
    'dtac': ('Subscriptions', '📱'),
    'ais': ('Subscriptions', '📱'),
    'advanced info service': ('Subscriptions', '📱'),
    'shopee thailand': ('Shopping', '👗'),
    'shopee pay': ('Shopping', '👗'),
    'lazada express': ('Shopping', '👗'),
    'payment to shopee': ('Shopping', '👗'),
    'grab holdings': ('Transport', '🚕'),
    'grab taxi': ('Transport', '🚕'),
    'bolt technology': ('Transport', '🚕'),
    'ptt public': ('Transport', '🚕'),
    'ptt oil': ('Transport', '🚕'),
    'bts group': ('Transport', '🚕'),
    'bangkok expressway': ('Transport', '🚕'),
    'metropolitan electricity': ('Housing', '🏠'),
    'metropolitan waterworks': ('Housing', '🏠'),
    'provincial electricity': ('Housing', '🏠'),
    'provincial waterworks': ('Housing', '🏠'),
    'pea': ('Housing', '🏠'),
    'mea': ('Housing', '🏠'),
    'bumrungrad': ('Health', '💊'),
    'samitivej': ('Health', '💊'),
    'bangkok dusit medical': ('Health', '💊'),
    'bdms': ('Health', '💊'),
    'boots retail': ('Health', '💊'),
    'watsons thailand': ('Health', '💊'),
    'google asia': ('Subscriptions', '📱'),
    'google payment': ('Subscriptions', '📱'),
    'apple itunes': ('Subscriptions', '📱'),
    'spotify': ('Subscriptions', '📱'),
    'netflix': ('Subscriptions', '📱'),

    # ─── ☕ Coffee ─────────────────────────────────────────────────────────
    'coffee': ('Coffee', '☕'),
    'latte': ('Coffee', '☕'),
    'cappuccino': ('Coffee', '☕'),
    'espresso': ('Coffee', '☕'),
    'americano': ('Coffee', '☕'),
    'mocha': ('Coffee', '☕'),
    'macchiato': ('Coffee', '☕'),
    'frappe': ('Coffee', '☕'),
    'starbucks': ('Coffee', '☕'),
    'cafe amazon': ('Coffee', '☕'),
    'amazon': ('Coffee', '☕'),
    'inthanin': ('Coffee', '☕'),
    'inthanin coffee': ('Coffee', '☕'),
    'wawee': ('Coffee', '☕'),
    'wawee coffee': ('Coffee', '☕'),
    'bluecup': ('Coffee', '☕'),
    'blue cup': ('Coffee', '☕'),
    'all cafe': ('Coffee', '☕'),
    'dd coffee': ('Coffee', '☕'),
    'pacamara': ('Coffee', '☕'),
    'roots': ('Coffee', '☕'),
    'roots coffee': ('Coffee', '☕'),
    'brave roasters': ('Coffee', '☕'),
    'kaizen coffee': ('Coffee', '☕'),
    'nana coffee': ('Coffee', '☕'),
    'ceresia': ('Coffee', '☕'),
    'factory coffee': ('Coffee', '☕'),
    'graph coffee': ('Coffee', '☕'),
    'akha ama': ('Coffee', '☕'),
    'doi chaang': ('Coffee', '☕'),
    'true coffee': ('Coffee', '☕'),
    'black canyon': ('Coffee', '☕'),
    'flat white': ('Coffee', '☕'),
    'cortado': ('Coffee', '☕'),
    'kopi': ('Coffee', '☕'),
    '% arabica': ('Coffee', '☕'),
    'ceresia': ('Coffee', '☕'),
    'graph coffee': ('Coffee', '☕'),
    'the coffee club': ('Coffee', '☕'),
    'coffee today': ('Coffee', '☕'),
    'tom n toms': ('Coffee', '☕'),
    'hollys': ('Coffee', '☕'),
    'costa coffee': ('Coffee', '☕'),
    'dean & deluca': ('Coffee', '☕'),
    'paul': ('Coffee', '☕'),
    'cafe': ('Coffee', '☕'),

    # ─── 🚕 Transport ─────────────────────────────────────────────────────
    'transport': ('Transport', '🚕'),
    'grab': ('Transport', '🚕'),
    'bolt': ('Transport', '🚕'),
    'bolt food': ('Food & Drinks', '🍜'),
    'grabfood': ('Food & Drinks', '🍜'),
    'grab food': ('Food & Drinks', '🍜'),
    'indrive': ('Transport', '🚕'),
    'in drive': ('Transport', '🚕'),
    'hailo': ('Transport', '🚕'),
    'maxim': ('Transport', '🚕'),
    'limousine': ('Transport', '🚕'),
    'taxi': ('Transport', '🚕'),
    'grab taxi': ('Transport', '🚕'),
    'ltc': ('Transport', '🚕'),
    'taxi meter': ('Transport', '🚕'),
    'grab car': ('Transport', '🚕'),
    'grab bike': ('Transport', '🚕'),
    'grabcar': ('Transport', '🚕'),
    'grabbike': ('Transport', '🚕'),
    'bolt': ('Transport', '🚕'),
    'bts': ('Transport', '🚕'),
    'mrt': ('Transport', '🚕'),
    'taxi': ('Transport', '🚕'),
    'motorbike': ('Transport', '🚕'),
    'motorbike taxi': ('Transport', '🚕'),
    'bike': ('Transport', '🚕'),
    'bus': ('Transport', '🚕'),
    'van': ('Transport', '🚕'),
    'minivan': ('Transport', '🚕'),
    'songthaew': ('Transport', '🚕'),
    'tuk tuk': ('Transport', '🚕'),
    'tuktuk': ('Transport', '🚕'),
    'boat': ('Transport', '🚕'),
    'ferry': ('Transport', '🚕'),
    'chao phraya': ('Transport', '🚕'),
    'airport link': ('Transport', '🚕'),
    'airport rail': ('Transport', '🚕'),
    'srt': ('Transport', '🚕'),
    'train': ('Transport', '🚕'),
    'parking': ('Transport', '🚕'),
    'toll': ('Transport', '🚕'),
    'expressway': ('Transport', '🚕'),
    'gas': ('Transport', '🚕'),
    'petrol': ('Transport', '🚕'),
    'gasoline': ('Transport', '🚕'),
    'ptt': ('Transport', '🚕'),
    'shell': ('Transport', '🚕'),
    'bangchak': ('Transport', '🚕'),
    'caltex': ('Transport', '🚕'),
    'esso': ('Transport', '🚕'),
    'indriver': ('Transport', '🚕'),
    'cabb': ('Transport', '🚕'),
    'ola': ('Transport', '🚕'),
    'van tour': ('Transport', '🚕'),
    'speed boat': ('Transport', '🚕'),
    'pier': ('Transport', '🚕'),
    'line taxi': ('Transport', '🚕'),
    'robinhood ride': ('Transport', '🚕'),

    # ─── 🛒 Groceries ─────────────────────────────────────────────────────
    'groceries': ('Groceries', '🛒'),
    'grocery': ('Groceries', '🛒'),
    'supermarket': ('Groceries', '🛒'),
    'market': ('Groceries', '🛒'),
    'fresh market': ('Groceries', '🛒'),
    'wet market': ('Groceries', '🛒'),
    'big c': ('Groceries', '🛒'),
    'bigc': ('Groceries', '🛒'),
    'tops': ('Groceries', '🛒'),
    'tops market': ('Groceries', '🛒'),
    'tops daily': ('Groceries', '🛒'),
    'lotus': ('Groceries', '🛒'),
    'tesco': ('Groceries', '🛒'),
    'tesco lotus': ('Groceries', '🛒'),
    'makro': ('Groceries', '🛒'),
    'villa market': ('Groceries', '🛒'),
    'villa': ('Groceries', '🛒'),
    'gourmet market': ('Groceries', '🛒'),
    'foodland': ('Groceries', '🛒'),
    'maxvalu': ('Groceries', '🛒'),
    'max value': ('Groceries', '🛒'),
    'donki': ('Groceries', '🛒'),
    'don don donki': ('Groceries', '🛒'),
    'don quijote': ('Groceries', '🛒'),
    'lawson': ('Groceries', '🛒'),
    'family mart': ('Groceries', '🛒'),
    'familymart': ('Groceries', '🛒'),
    '7-eleven': ('Groceries', '🛒'),
    '7-11': ('Groceries', '🛒'),
    '711': ('Groceries', '🛒'),
    'seven eleven': ('Groceries', '🛒'),
    'mini big c': ('Groceries', '🛒'),
    'cp fresh mart': ('Groceries', '🛒'),
    'cp freshmart': ('Groceries', '🛒'),
    'jiffy': ('Groceries', '🛒'),
    'donki': ('Groceries', '🛒'),
    'don don donki': ('Groceries', '🛒'),
    'fresh mart': ('Groceries', '🛒'),
    'talad': ('Groceries', '🛒'),
    'ampol food': ('Groceries', '🛒'),
    'big c extra': ('Groceries', '🛒'),
    'lotus express': ('Groceries', '🛒'),
    'the mall': ('Groceries', '🛒'),
    'eathai': ('Groceries', '🛒'),
    'rimping': ('Groceries', '🛒'),

    # ─── 🏠 Housing ───────────────────────────────────────────────────────
    'housing': ('Housing', '🏠'),
    'rent': ('Housing', '🏠'),
    'condo': ('Housing', '🏠'),
    'apartment': ('Housing', '🏠'),
    'utilities': ('Housing', '🏠'),
    'electric': ('Housing', '🏠'),
    'electricity': ('Housing', '🏠'),
    'electric bill': ('Housing', '🏠'),
    'pea': ('Housing', '🏠'),
    'mea': ('Housing', '🏠'),
    'water bill': ('Housing', '🏠'),
    'mwa': ('Housing', '🏠'),
    'internet': ('Housing', '🏠'),
    'wifi': ('Housing', '🏠'),
    'true internet': ('Housing', '🏠'),
    'ais fibre': ('Housing', '🏠'),
    '3bb': ('Housing', '🏠'),
    'tot': ('Housing', '🏠'),
    'nt': ('Housing', '🏠'),
    'phone bill': ('Housing', '🏠'),
    'mobile bill': ('Housing', '🏠'),
    'ais': ('Housing', '🏠'),
    'dtac': ('Housing', '🏠'),
    'true move': ('Housing', '🏠'),
    'truemove': ('Housing', '🏠'),
    'laundry': ('Housing', '🏠'),
    'cleaning': ('Housing', '🏠'),
    'maid': ('Housing', '🏠'),
    'furniture': ('Housing', '🏠'),
    'ikea': ('Housing', '🏠'),
    'home pro': ('Housing', '🏠'),
    'homepro': ('Housing', '🏠'),
    'thai watsadu': ('Housing', '🏠'),
    'baan & beyond': ('Housing', '🏠'),
    'index living': ('Housing', '🏠'),
    'sb furniture': ('Housing', '🏠'),
    'mr diy': ('Housing', '🏠'),
    'aircon': ('Housing', '🏠'),
    'air conditioner': ('Housing', '🏠'),
    'air con': ('Housing', '🏠'),
    'true online': ('Housing', '🏠'),
    'ais broadband': ('Housing', '🏠'),
    'pest control': ('Housing', '🏠'),
    'internet bill': ('Housing', '🏠'),

    # ─── 💊 Health ────────────────────────────────────────────────────────
    'health': ('Health', '💊'),
    'pharmacy': ('Health', '💊'),
    'clinic': ('Health', '💊'),
    'hospital': ('Health', '💊'),
    'doctor': ('Health', '💊'),
    'dentist': ('Health', '💊'),
    'dental': ('Health', '💊'),
    'medicine': ('Health', '💊'),
    'drug': ('Health', '💊'),
    'drugs': ('Health', '💊'),
    'vitamin': ('Health', '💊'),
    'vitamins': ('Health', '💊'),
    'multivitamin': ('Health', '💊'),
    'supplement': ('Health', '💊'),
    'supplements': ('Health', '💊'),
    # Brands
    'swisse': ('Health', '💊'),
    'dr pong': ('Health', '💊'),
    'dr.pong': ('Health', '💊'),
    'blackmores': ('Health', '💊'),
    'centrum': ('Health', '💊'),
    'vistra': ('Health', '💊'),
    'mega we care': ('Health', '💊'),
    'mega wecare': ('Health', '💊'),
    'amsel': ('Health', '💊'),
    'bioganic': ('Health', '💊'),
    'real elixir': ('Health', '💊'),
    'dhc': ('Health', '💊'),
    'nature bounty': ('Health', '💊'),
    'gnc': ('Health', '💊'),
    'now foods': ('Health', '💊'),
    'solgar': ('Health', '💊'),
    'whey protein': ('Health', '💊'),
    'protein': ('Health', '💊'),
    'collagen': ('Health', '💊'),
    'probiotic': ('Health', '💊'),
    'fish oil': ('Health', '💊'),
    'omega': ('Health', '💊'),
    'zinc': ('Health', '💊'),
    'magnesium': ('Health', '💊'),
    # Pharmacies & hospitals
    'boots': ('Health', '💊'),
    'watsons': ('Health', '💊'),
    'fascino': ('Health', '💊'),
    'bumrungrad': ('Health', '💊'),
    'samitivej': ('Health', '💊'),
    'bangkok hospital': ('Health', '💊'),
    'medpark': ('Health', '💊'),
    'praram 9': ('Health', '💊'),
    'phyathai': ('Health', '💊'),
    'dr.pong': ('Health', '💊'),
    'mega we care': ('Health', '💊'),
    'amsel': ('Health', '💊'),
    'dhc': ('Health', '💊'),
    'gnc': ('Health', '💊'),
    'now foods': ('Health', '💊'),
    'nature bounty': ('Health', '💊'),
    'real elixir': ('Health', '💊'),
    'bioganic': ('Health', '💊'),
    'protein': ('Health', '💊'),
    'whey protein': ('Health', '💊'),
    'whey': ('Health', '💊'),
    'creatine': ('Health', '💊'),
    'bcaa': ('Health', '💊'),
    'collagen': ('Health', '💊'),
    'probiotic': ('Health', '💊'),
    'probiotics': ('Health', '💊'),
    'fish oil': ('Health', '💊'),
    'omega 3': ('Health', '💊'),
    'zinc': ('Health', '💊'),
    'magnesium': ('Health', '💊'),
    'melatonin': ('Health', '💊'),
    'sunscreen': ('Health', '💊'),
    'spf': ('Health', '💊'),
    'face wash': ('Health', '💊'),
    'shampoo': ('Health', '💊'),
    'toothpaste': ('Health', '💊'),
    'toothbrush': ('Health', '💊'),
    'electric toothbrush': ('Health', '💊'),
    'panadol': ('Health', '💊'),
    'tylenol': ('Health', '💊'),
    'ibuprofen': ('Health', '💊'),
    'antigen test': ('Health', '💊'),
    'rapid test': ('Health', '💊'),
    'paolo': ('Health', '💊'),
    'sikarin': ('Health', '💊'),
    'bnh': ('Health', '💊'),
    'eye care': ('Health', '💊'),
    'glasses': ('Health', '💊'),
    'contact lens': ('Health', '💊'),
    'sunscreen': ('Health', '💊'),
    'skincare': ('Health', '💊'),

    # ─── 👗 Shopping ──────────────────────────────────────────────────────
    'shopping': ('Shopping', '👗'),
    'clothes': ('Shopping', '👗'),
    'clothing': ('Shopping', '👗'),
    'accessories': ('Shopping', '👗'),
    'shoes': ('Shopping', '👗'),
    'sneakers': ('Shopping', '👗'),
    'bag': ('Shopping', '👗'),
    'bags': ('Shopping', '👗'),
    'watch': ('Shopping', '👗'),
    'jewelry': ('Shopping', '👗'),
    'perfume': ('Shopping', '👗'),
    'cosmetics': ('Shopping', '👗'),
    'makeup': ('Shopping', '👗'),
    # Brands & stores
    'uniqlo': ('Shopping', '👗'),
    'h&m': ('Shopping', '👗'),
    'zara': ('Shopping', '👗'),
    'muji': ('Shopping', '👗'),
    'cotton on': ('Shopping', '👗'),
    'gu': ('Shopping', '👗'),
    'nike': ('Shopping', '👗'),
    'adidas': ('Shopping', '👗'),
    'converse': ('Shopping', '👗'),
    'vans': ('Shopping', '👗'),
    'new balance': ('Shopping', '👗'),
    'charles & keith': ('Shopping', '👗'),
    'aldo': ('Shopping', '👗'),
    'sephora': ('Shopping', '👗'),
    'eveandboy': ('Shopping', '👗'),
    'eve and boy': ('Shopping', '👗'),
    'beautrium': ('Shopping', '👗'),
    'naraya': ('Shopping', '👗'),
    'jaspal': ('Shopping', '👗'),
    'cc double o': ('Shopping', '👗'),
    'gentlewoman': ('Shopping', '👗'),
    'pomelo': ('Shopping', '👗'),
    # Online shopping
    'shopee': ('Shopping', '👗'),
    'lazada': ('Shopping', '👗'),
    'shopeepay': ('Shopping', '👗'),
    'tiktok shop': ('Shopping', '👗'),
    'amazon': ('Shopping', '👗'),
    'aliexpress': ('Shopping', '👗'),
    # Tech & electronics
    'headphones': ('Shopping', '👗'),
    'earbuds': ('Shopping', '👗'),
    'airpods': ('Shopping', '👗'),
    'sennheiser': ('Shopping', '👗'),
    'sony': ('Shopping', '👗'),
    'samsung': ('Shopping', '👗'),
    'apple': ('Shopping', '👗'),
    'iphone': ('Shopping', '👗'),
    'ipad': ('Shopping', '👗'),
    'macbook': ('Shopping', '👗'),
    'laptop': ('Shopping', '👗'),
    'phone case': ('Shopping', '👗'),
    'charger': ('Shopping', '👗'),
    'power bank': ('Shopping', '👗'),
    'banana it': ('Shopping', '👗'),
    'jib': ('Shopping', '👗'),
    'power buy': ('Shopping', '👗'),
    'powerbuy': ('Shopping', '👗'),
    'daiso': ('Shopping', '👗'),
    'miniso': ('Shopping', '👗'),
    'loft': ('Shopping', '👗'),
    'sephora': ('Shopping', '👗'),
    'eveandboy': ('Shopping', '👗'),
    'eve and boy': ('Shopping', '👗'),
    'beautrium': ('Shopping', '👗'),
    'mac cosmetics': ('Shopping', '👗'),
    'innisfree': ('Shopping', '👗'),
    'the face shop': ('Shopping', '👗'),
    'etude': ('Shopping', '👗'),
    'missha': ('Shopping', '👗'),
    'it city': ('Shopping', '👗'),
    'studio 7': ('Shopping', '👗'),
    'istudio': ('Shopping', '👗'),
    # Malls
    'central': ('Shopping', '👗'),
    'centralworld': ('Shopping', '👗'),
    'siam paragon': ('Shopping', '👗'),
    'emquartier': ('Shopping', '👗'),
    'emporium': ('Shopping', '👗'),
    'iconsiam': ('Shopping', '👗'),
    'terminal 21': ('Shopping', '👗'),
    'mbk': ('Shopping', '👗'),
    'platinum': ('Shopping', '👗'),
    'pratunam': ('Shopping', '👗'),
    'chatuchak': ('Shopping', '👗'),
    'jj market': ('Shopping', '👗'),
    'robinson': ('Shopping', '👗'),
    'mega bangna': ('Shopping', '👗'),
    'fashion island': ('Shopping', '👗'),
    'future park': ('Shopping', '👗'),
    'seacon': ('Shopping', '👗'),

    # ─── 🎉 Entertainment ─────────────────────────────────────────────────
    'entertainment': ('Entertainment', '🎉'),
    'movies': ('Entertainment', '🎉'),
    'movie': ('Entertainment', '🎉'),
    'cinema': ('Entertainment', '🎉'),
    'sf cinema': ('Entertainment', '🎉'),
    'major cineplex': ('Entertainment', '🎉'),
    'major': ('Entertainment', '🎉'),
    'concert': ('Entertainment', '🎉'),
    'night out': ('Entertainment', '🎉'),
    'bar': ('Entertainment', '🎉'),
    'pub': ('Entertainment', '🎉'),
    'club': ('Entertainment', '🎉'),
    'nightclub': ('Entertainment', '🎉'),
    'karaoke': ('Entertainment', '🎉'),
    'ktv': ('Entertainment', '🎉'),
    'bowling': ('Entertainment', '🎉'),
    'arcade': ('Entertainment', '🎉'),
    'escape room': ('Entertainment', '🎉'),
    'museum': ('Entertainment', '🎉'),
    'exhibition': ('Entertainment', '🎉'),
    'art gallery': ('Entertainment', '🎉'),
    'zoo': ('Entertainment', '🎉'),
    'safari world': ('Entertainment', '🎉'),
    'siam ocean world': ('Entertainment', '🎉'),
    'madame tussauds': ('Entertainment', '🎉'),
    'kidzania': ('Entertainment', '🎉'),
    'theme park': ('Entertainment', '🎉'),
    'water park': ('Entertainment', '🎉'),
    'spa': ('Entertainment', '🎉'),
    'massage': ('Entertainment', '🎉'),
    'thai massage': ('Entertainment', '🎉'),
    'onsen': ('Entertainment', '🎉'),
    'let relax': ('Entertainment', '🎉'),
    'health land': ('Entertainment', '🎉'),
    'gym': ('Entertainment', '🎉'),
    'fitness': ('Entertainment', '🎉'),
    'fitness first': ('Entertainment', '🎉'),
    'virgin active': ('Entertainment', '🎉'),
    'jetts': ('Entertainment', '🎉'),
    'anytime fitness': ('Entertainment', '🎉'),
    'yoga': ('Entertainment', '🎉'),
    'swimming': ('Entertainment', '🎉'),
    'pool': ('Entertainment', '🎉'),
    'game': ('Entertainment', '🎉'),
    'games': ('Entertainment', '🎉'),
    'steam': ('Entertainment', '🎉'),
    'playstation': ('Entertainment', '🎉'),
    'nintendo': ('Entertainment', '🎉'),
    'hbo': ('Entertainment', '🎉'),
    'hbo go': ('Entertainment', '🎉'),
    'wetv': ('Entertainment', '🎉'),
    'viu': ('Entertainment', '🎉'),
    'monomax': ('Entertainment', '🎉'),
    'true id': ('Entertainment', '🎉'),
    'joox': ('Entertainment', '🎉'),
    'muay thai': ('Entertainment', '🎉'),
    'boxing': ('Entertainment', '🎉'),
    'pilates': ('Entertainment', '🎉'),
    'crossfit': ('Entertainment', '🎉'),
    'badminton': ('Entertainment', '🎉'),
    'tennis': ('Entertainment', '🎉'),
    'golf': ('Entertainment', '🎉'),
    'running': ('Entertainment', '🎉'),
    'beer': ('Entertainment', '🎉'),
    'wine': ('Entertainment', '🎉'),
    'whiskey': ('Entertainment', '🎉'),
    'cocktail': ('Entertainment', '🎉'),
    'alcohol': ('Entertainment', '🎉'),
    'singha': ('Entertainment', '🎉'),
    'chang': ('Entertainment', '🎉'),
    'leo': ('Entertainment', '🎉'),
    'heineken': ('Entertainment', '🎉'),

    # ─── 📱 Subscriptions ─────────────────────────────────────────────────
    'subscription': ('Subscriptions', '📱'),
    'subscriptions': ('Subscriptions', '📱'),
    'youtube': ('Subscriptions', '📱'),
    'youtube premium': ('Subscriptions', '📱'),
    'netflix': ('Subscriptions', '📱'),
    'spotify': ('Subscriptions', '📱'),
    'apple music': ('Subscriptions', '📱'),
    'disney plus': ('Subscriptions', '📱'),
    'disney+': ('Subscriptions', '📱'),
    'hbo go': ('Subscriptions', '📱'),
    'viu': ('Subscriptions', '📱'),
    'wetv': ('Subscriptions', '📱'),
    'iqiyi': ('Subscriptions', '📱'),
    'google one': ('Subscriptions', '📱'),
    'icloud': ('Subscriptions', '📱'),
    'chatgpt': ('Subscriptions', '📱'),
    'openai': ('Subscriptions', '📱'),
    'canva': ('Subscriptions', '📱'),
    'adobe': ('Subscriptions', '📱'),
    'notion': ('Subscriptions', '📱'),
    'line': ('Subscriptions', '📱'),
    'true id': ('Subscriptions', '📱'),
    'ais play': ('Subscriptions', '📱'),
    'monomax': ('Subscriptions', '📱'),

    # ─── ✈️ Travel ─────────────────────────────────────────────────────────
    'travel': ('Travel', '✈️'),
    'hotel': ('Travel', '✈️'),
    'hostel': ('Travel', '✈️'),
    'airbnb': ('Travel', '✈️'),
    'agoda': ('Travel', '✈️'),
    'booking.com': ('Travel', '✈️'),
    'flight': ('Travel', '✈️'),
    'airline': ('Travel', '✈️'),
    'airasia': ('Travel', '✈️'),
    'nok air': ('Travel', '✈️'),
    'lion air': ('Travel', '✈️'),
    'thai airways': ('Travel', '✈️'),
    'vietjet': ('Travel', '✈️'),
    'bangkok airways': ('Travel', '✈️'),
    'trip': ('Travel', '✈️'),
    'vacation': ('Travel', '✈️'),
    'holiday': ('Travel', '✈️'),
    'passport': ('Travel', '✈️'),
    'visa fee': ('Travel', '✈️'),
    'luggage': ('Travel', '✈️'),
    'suitcase': ('Travel', '✈️'),
    'klook': ('Travel', '✈️'),
    'kkday': ('Travel', '✈️'),
    'traveloka': ('Travel', '✈️'),

    # ─── 🎓 School ────────────────────────────────────────────────────────
    'school': ('School', '🎓'),
    'tuition': ('School', '🎓'),
    'books': ('School', '🎓'),
    'book': ('School', '🎓'),
    'textbook': ('School', '🎓'),
    'stationery': ('School', '🎓'),
    'language': ('School', '🎓'),
    'italian': ('School', '🎓'),
    'class': ('School', '🎓'),
    'course': ('School', '🎓'),
    'workshop': ('School', '🎓'),
    'seminar': ('School', '🎓'),
    'training': ('School', '🎓'),
    'udemy': ('School', '🎓'),
    'skillshare': ('School', '🎓'),
    'coursera': ('School', '🎓'),
    'se-ed': ('School', '🎓'),
    'b2s': ('School', '🎓'),
    'kinokuniya': ('School', '🎓'),
    'asia books': ('School', '🎓'),

    # ─── 🚬 Cigarettes ──────────────────────────────────────────────────
    'cigarette': ('Cigarettes', '🚬'),
    'cigarettes': ('Cigarettes', '🚬'),
    'cig': ('Cigarettes', '🚬'),
    'cigs': ('Cigarettes', '🚬'),
    'smoke': ('Cigarettes', '🚬'),
    'smoking': ('Cigarettes', '🚬'),
    'tobacco': ('Cigarettes', '🚬'),
    'marlboro': ('Cigarettes', '🚬'),
    'lm': ('Cigarettes', '🚬'),
    'winston': ('Cigarettes', '🚬'),
    'camel': ('Cigarettes', '🚬'),
    'lucky strike': ('Cigarettes', '🚬'),
    'dunhill': ('Cigarettes', '🚬'),
    'esse': ('Cigarettes', '🚬'),
    'krong thip': ('Cigarettes', '🚬'),
    'krongthip': ('Cigarettes', '🚬'),
    'sai fon': ('Cigarettes', '🚬'),
    'falling rain': ('Cigarettes', '🚬'),
    'wonder': ('Cigarettes', '🚬'),
    'vape': ('Cigarettes', '🚬'),
    'pod': ('Cigarettes', '🚬'),
    'iqos': ('Cigarettes', '🚬'),
    'relx': ('Cigarettes', '🚬'),

    # ─── 💵 Income ────────────────────────────────────────────────────────
    'salary': ('Salary', '💵'),
    'freelance': ('Freelance', '💼'),
    'gallery': ('Gallery Sales', '🖼️'),
    'gallery sales': ('Gallery Sales', '🖼️'),
    'artwork': ('Artwork / Commission', '🎨'),
    'sold': ('Gallery Sales', '🖼️'),
    'commission': ('Artwork / Commission', '🎨'),
    'bonus': ('Bonus', '🏆'),
    'refund': ('Cashback / Refund', '💳'),
    'cashback': ('Cashback / Refund', '💳'),
    'dividend': ('Investment', '💹'),
    'interest': ('Investment', '💹'),
    'investment': ('Investment', '💹'),
    'gift money': ('Gift Money', '🎁'),
    'angpao': ('Gift Money', '🎁'),
    'prize': ('Bonus', '🏆'),
    'business': ('Business', '🤝'),
}

CATEGORY_LIST = [
    ('🍜', 'Food & Drinks'),
    ('☕', 'Coffee'),
    ('🚕', 'Transport'),
    ('🛒', 'Groceries'),
    ('🏠', 'Housing'),
    ('💊', 'Health'),
    ('👗', 'Shopping'),
    ('🎉', 'Entertainment'),
    ('📱', 'Subscriptions'),
    ('✈️', 'Travel'),
    ('🎓', 'School'),
    ('🚬', 'Cigarettes'),
    ('🧾', 'Other'),
]

INCOME_CATEGORY_LIST = [
    ('💵', 'Salary'),
    ('💼', 'Freelance'),
    ('🖼️', 'Gallery Sales'),
    ('🎨', 'Artwork / Commission'),
    ('🏆', 'Bonus'),
    ('🎁', 'Gift Money'),
    ('💳', 'Cashback / Refund'),
    ('💹', 'Investment'),
    ('🤝', 'Business'),
    ('🧾', 'Other Income'),
]

ACCOUNT_KEYWORDS = {
    'bank': 'Bangkok Bank',
    'bangkok bank': 'Bangkok Bank',
    'bbl': 'Bangkok Bank',
    'true money': 'True Money Wallet',
    'truemoney': 'True Money Wallet',
    'true wallet': 'True Money Wallet',
    'mrt': 'MRT EMV Visa',
    'emv': 'MRT EMV Visa',
    'visa': 'MRT EMV Visa',
    'rabbit': 'Rabbit Card',
    'rabbit card': 'Rabbit Card',
    'cash': 'Cash',
    'muvmi': 'Muvmi',
    'solsot': 'Solsot Member',
    'solsot member': 'Solsot Member',
}


# ─── Database ─────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        balance REAL NOT NULL DEFAULT 0.0,
        UNIQUE(user_id, name)
    );
    CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        emoji TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        description TEXT,
        type TEXT NOT NULL,
        category TEXT NOT NULL DEFAULT 'Other',
        account TEXT NOT NULL DEFAULT 'Cash',
        timestamp DATETIME DEFAULT (datetime('now', '+7 hours'))
    );
    CREATE TABLE IF NOT EXISTS recurring_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        amount REAL NOT NULL,
        category TEXT NOT NULL DEFAULT 'Subscriptions',
        account TEXT NOT NULL DEFAULT 'Cash',
        next_due_date DATE NOT NULL,
        frequency TEXT NOT NULL DEFAULT 'monthly'
    );
    """)
    for emoji, name in CATEGORY_LIST:
        c.execute("INSERT OR IGNORE INTO categories (name, emoji) VALUES (?, ?)", (name, emoji))
    # Create accounts if they don't exist (with 0 balance — use /updatebalance to set correct values)
    # INSERT OR IGNORE means this only runs on a fresh database, never overwrites existing balances
    uid = AUTHORIZED_USER_ID
    for user_id, name in [
        (uid, 'Bangkok Bank'),
        (uid, 'True Money Wallet'),
        (uid, 'MRT EMV Visa'),
        (uid, 'Rabbit Card'),
        (uid, 'Cash'),
        (uid, 'Muvmi'),
        (uid, 'Solsot Member'),
    ]:
        c.execute("INSERT OR IGNORE INTO accounts (user_id, name, balance) VALUES (?, ?, 0.0)", (user_id, name))
    # Note: Subscriptions are managed via /addsubscription and /deletesubscription commands.
    # No hardcoded subscription seeds — the live database on Render has the real data.
    conn.commit()
    conn.close()


def get_db():
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    return conn


# ─── Helpers ──────────────────────────────────────────────────────────────
def detect_category(text):
    text_lower = text.lower()
    sorted_keywords = sorted(CATEGORIES.keys(), key=len, reverse=True)
    for keyword in sorted_keywords:
        if keyword in text_lower:
            return CATEGORIES[keyword]
    return ('Other', '🧾')


def detect_account(text):
    text_lower = text.lower()
    sorted_keywords = sorted(ACCOUNT_KEYWORDS.keys(), key=len, reverse=True)
    for keyword in sorted_keywords:
        if keyword in text_lower:
            return ACCOUNT_KEYWORDS[keyword]
    return 'Cash'


# ─── Recurring subscription auto-processing ─────────────────────────────────
def process_due_subscriptions():
    """Check and process any overdue subscriptions.
    Returns list of (sub_name, amount, account, user_id) tuples that were processed.
    """
    conn = get_db()
    c = conn.cursor()
    today = datetime.now(BANGKOK_TZ).date()
    today_str = str(today)

    logger.info(f"Checking subscriptions for date: {today_str}")

    c.execute(
        "SELECT id, user_id, name, amount, category, account, next_due_date, frequency "
        "FROM recurring_subscriptions WHERE next_due_date <= ?",
        (today_str,)
    )
    due_subs = c.fetchall()
    processed = []

    logger.info(f"Found {len(due_subs)} due subscription(s)")

    for sub in due_subs:
        amount = sub['amount']
        logger.info(f"Processing subscription: {sub['name']} ฿{amount} from {sub['account']} (due: {sub['next_due_date']})")

        c.execute(
            "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'expense', ?, ?)",
            (sub['user_id'], -amount, f"🔄 {sub['name']} (auto)", sub['category'], sub['account'])
        )
        c.execute(
            "UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?",
            (amount, sub['user_id'], sub['account'])
        )

        # Calculate next due date
        due_date = datetime.strptime(sub['next_due_date'], "%Y-%m-%d").date()
        if sub['frequency'] == 'monthly':
            if due_date.month == 12:
                next_due = due_date.replace(year=due_date.year + 1, month=1)
            else:
                try:
                    next_due = due_date.replace(month=due_date.month + 1)
                except ValueError:
                    next_due = (due_date.replace(day=1) + timedelta(days=32)).replace(day=due_date.day)
        elif sub['frequency'] == 'weekly':
            next_due = due_date + timedelta(weeks=1)
        elif sub['frequency'] == 'yearly':
            next_due = due_date.replace(year=due_date.year + 1)
        else:
            next_due = due_date + timedelta(days=30)

        c.execute("UPDATE recurring_subscriptions SET next_due_date = ? WHERE id = ?", (str(next_due), sub['id']))
        processed.append((sub['name'], amount, sub['account'], sub['user_id']))
        logger.info(f"Subscription '{sub['name']}' logged. Next due: {next_due}")

    conn.commit()
    conn.close()
    return processed


# ─── Google Drive Monthly Backup ─────────────────────────────────────────────
GDRIVE_FOLDER_ID = '1_APGayoUno9u-7-nKHZ-Fxk0ntt-WRqg'


def get_gdrive_service():
    """Build Google Drive service from service account JSON in env var."""
    creds_json = os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', '')
    if not creds_json:
        logger.error("GOOGLE_SERVICE_ACCOUNT_JSON not set in environment.")
        return None
    try:
        creds_dict = json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/drive.file']
        )
        return build('drive', 'v3', credentials=creds)
    except Exception as e:
        logger.error(f"Failed to build Drive service: {e}")
        return None


def upload_to_gdrive(file_bytes: bytes, filename: str, error_out: list = None) -> str:
    """Upload a file to Google Drive. Returns the file URL or empty string.
    If error_out is provided (a list), the real exception text is appended
    to it on failure so callers can surface Google's actual error message."""
    if not GDRIVE_AVAILABLE:
        logger.error("Google API packages not available.")
        if error_out is not None:
            error_out.append("Google API packages not available.")
        return ''
    service = get_gdrive_service()
    if not service:
        if error_out is not None:
            error_out.append("Could not build Google Drive service — check GOOGLE_SERVICE_ACCOUNT_JSON is valid.")
        return ''
    try:
        file_metadata = {
            'name': filename,
            'parents': [GDRIVE_FOLDER_ID]
        }
        media = MediaIoBaseUpload(
            io.BytesIO(file_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=False
        )
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, webViewLink'
        ).execute()
        link = uploaded.get('webViewLink', '')
        logger.info(f"Uploaded {filename} to Google Drive: {link}")
        return link
    except Exception as e:
        logger.error(f"Google Drive upload failed: {e}")
        if error_out is not None:
            error_out.append(str(e))
        return ''


def send_excel_email(file_bytes: bytes, filename: str, subject: str, body: str) -> str:
    """
    Email an Excel file as an attachment via the Resend HTTPS API.
    (Switched from Gmail SMTP because Render blocks outbound SMTP ports
    25/465/587 on free-tier web services — Resend uses HTTPS instead,
    which is never blocked.)
    Returns '' (empty string) on success, or an error message string on failure.
    Requires these Render env vars:
      RESEND_API_KEY   — from resend.com → API Keys
      BACKUP_EMAIL_TO  — the address to send the backup to. On Resend's free
                          tier (no verified custom domain), this MUST be the
                          same email address you used to sign up for Resend.
    """
    api_key = os.environ.get('RESEND_API_KEY', '').strip()
    to_email = os.environ.get('BACKUP_EMAIL_TO', '').strip()

    if not api_key:
        return 'RESEND_API_KEY environment variable is not set on Render.'
    if not to_email:
        return 'BACKUP_EMAIL_TO environment variable is not set on Render (this must be the same email you signed up to Resend with, unless you have a verified domain).'

    try:
        payload = {
            'from': 'Axe Finance <onboarding@resend.dev>',
            'to': [to_email],
            'subject': subject,
            'text': body,
            'attachments': [
                {
                    'filename': filename,
                    'content': base64.b64encode(file_bytes).decode('ascii'),
                }
            ],
        }
        resp = requests.post(
            'https://api.resend.com/emails',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            json=payload,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            logger.info(f"Emailed backup '{filename}' to {to_email} via Resend")
            return ''
        else:
            error_detail = resp.text
            logger.error(f"Resend email failed ({resp.status_code}): {error_detail}")
            return f"Resend API returned {resp.status_code}: {error_detail}"
    except Exception as e:
        logger.error(f"Email backup failed: {e}")
        return str(e)


def generate_monthly_excel(year: int, month: int) -> bytes:
    """Generate a polished, multi-sheet Excel report for a given month,
    styled to match the Axe Finance light-theme dashboard (white cards,
    sapphire-blue accent, green/red semantic colors). Produces three sheets:
    Summary, Transactions, and By Category. Returns raw .xlsx bytes."""
    import calendar as cal_mod

    month_start = f"{year}-{month:02d}-01"
    month_end   = f"{year+1}-01-01" if month == 12 else f"{year}-{month+1:02d}-01"

    conn = get_db()
    c    = conn.cursor()
    c.execute(
        "SELECT timestamp, type, amount, description, category, account "
        "FROM transactions WHERE user_id = ? AND timestamp >= ? AND timestamp < ? "
        "ORDER BY timestamp",
        (AUTHORIZED_USER_ID, month_start, month_end)
    )
    txns = c.fetchall()
    c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (AUTHORIZED_USER_ID,))
    accounts = c.fetchall()
    conn.close()

    label   = f"{cal_mod.month_name[month]} {year}"
    now_str = datetime.now(BANGKOK_TZ).strftime('%d %b %Y %H:%M')

    # ── Palette — mirrors the Axe Finance light-theme dashboard exactly ────
    INK      = "1D1D1F"   # --text / --nav-bg
    INK_SOFT = "6E6E73"   # --text2
    PANEL    = "F5F5F7"   # --bg / --badge-bg
    PANEL2   = "FAFAFA"   # --bg3
    WHITE    = "FFFFFF"
    BORDER_C = "E5E5EA"
    ACCENT   = "3763A8"   # --gold (sapphire accent)
    GREEN    = "34C759"   # --up
    GREEN_BG = "EBF9EE"   # --up-bg
    RED      = "FF3B30"   # --down
    RED_BG   = "FFEBEA"   # --down-bg

    money_fmt = '#,##0.00'
    pct_fmt   = '0.0%'
    center    = Alignment(horizontal="center", vertical="center")
    left_a    = Alignment(horizontal="left",   vertical="center")
    right_a   = Alignment(horizontal="right",  vertical="center")

    def fill(hexcolor):
        return PatternFill("solid", fgColor=hexcolor)

    def thin_border(color=BORDER_C):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def header_border():
        s = Side(style="thin", color=BORDER_C)
        return Border(left=s, right=s, top=s, bottom=Side(style="medium", color=ACCENT))

    def f_cell(bold=False, color=INK, size=10.5):
        return Font(name="Calibri", bold=bold, size=size, color=color)

    def f_header():
        return Font(name="Calibri", bold=True, size=10.5, color=ACCENT)

    def title_banner(ws, text, subtitle, span_cols):
        last_col = get_column_letter(span_cols)
        ws.merge_cells(f"A1:{last_col}1")
        t = ws["A1"]
        t.value = text
        t.font = Font(name="Calibri", bold=True, size=18, color=ACCENT)
        t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f"A2:{last_col}2")
        s = ws["A2"]
        s.value = subtitle
        s.font = Font(name="Calibri", size=9.5, color=INK_SOFT)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        s.border = Border(bottom=Side(style="medium", color=ACCENT))
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 8

    # Totals used across sheets
    income_total = 0.0
    expense_total = 0.0
    for t in txns:
        amt = float(t['amount'])
        if t['type'] == 'income':
            income_total += amt
        elif t['type'] == 'expense':
            expense_total += abs(amt)
    net = income_total - expense_total
    savings_rate = (net / income_total) if income_total > 0 else 0.0
    total_balance = sum(float(a['balance']) for a in accounts)

    wb = Workbook()

    # ═══════════════════════ Sheet 1 — Summary ════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = ACCENT
    title_banner(ws1, f"Axe Finance  —  {label}",
                 f"Generated {now_str} (Bangkok Time)  ·  Amounts in Thai Baht (฿)", 2)

    r = 5
    ws1.merge_cells(f"A{r}:B{r}")
    hero_lbl = ws1.cell(r, 1, "NET WORTH (CURRENT)")
    hero_lbl.font = Font(name="Calibri", bold=True, size=10, color=INK_SOFT)
    hero_lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[r].height = 22
    r += 1
    ws1.merge_cells(f"A{r}:B{r}")
    hero_val = ws1.cell(r, 1, total_balance)
    hero_val.number_format = money_fmt
    hero_val.font = Font(name="Calibri", bold=True, size=26, color=ACCENT)
    hero_val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hero_val.border = Border(bottom=Side(style="medium", color=ACCENT))
    ws1.row_dimensions[r].height = 40
    r += 2

    def kpi_row(row, label_text, value, color, value_bg, num_fmt):
        lc = ws1.cell(row, 1, label_text)
        lc.font = f_cell(color=INK_SOFT, size=10.5)
        lc.fill = fill(PANEL)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = Border(left=Side(style="thick", color=color),
                            top=Side(style="thin", color=BORDER_C),
                            bottom=Side(style="thin", color=BORDER_C))
        vc = ws1.cell(row, 2, value)
        vc.number_format = num_fmt
        vc.font = f_cell(bold=True, color=color, size=13)
        vc.fill = fill(value_bg)
        vc.alignment = right_a
        vc.border = Border(right=Side(style="thin", color=BORDER_C),
                            top=Side(style="thin", color=BORDER_C),
                            bottom=Side(style="thin", color=BORDER_C))
        ws1.row_dimensions[row].height = 28

    sec1 = ws1.cell(r, 1, f"FOR {cal_mod.month_name[month].upper()} {year}")
    sec1.font = Font(name="Calibri", bold=True, size=9.5, color=INK_SOFT)
    r += 1
    kpi_row(r, "Total Income", income_total, GREEN, GREEN_BG, money_fmt); r += 1
    kpi_row(r, "Total Expenses", expense_total, RED, RED_BG, money_fmt); r += 1
    kpi_row(r, "Net Savings", net, GREEN if net >= 0 else RED,
            GREEN_BG if net >= 0 else RED_BG, money_fmt); r += 1
    kpi_row(r, "Savings Rate", savings_rate, GREEN if savings_rate >= 0 else RED,
            GREEN_BG if savings_rate >= 0 else RED_BG, pct_fmt); r += 2

    sec2 = ws1.cell(r, 1, "ACCOUNT BALANCES")
    sec2.font = Font(name="Calibri", bold=True, size=9.5, color=INK_SOFT)
    r += 1
    for acc in accounts:
        name, bal = acc['name'], float(acc['balance'])
        nc = ws1.cell(r, 1, name)
        nc.font = f_cell(color=INK)
        nc.fill = fill(WHITE)
        nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        nc.border = thin_border()
        bc = ws1.cell(r, 2, bal)
        bc.number_format = money_fmt
        bc.font = f_cell(bold=True, color=ACCENT)
        bc.fill = fill(WHITE)
        bc.alignment = right_a
        bc.border = thin_border()
        ws1.row_dimensions[r].height = 22
        r += 1

    tc = ws1.cell(r, 1, "Total")
    tc.font = f_cell(bold=True, color=INK)
    tc.fill = fill(PANEL)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tc.border = thin_border()
    tbc = ws1.cell(r, 2, total_balance)
    tbc.number_format = money_fmt
    tbc.font = f_cell(bold=True, color=INK, size=12)
    tbc.fill = fill(PANEL)
    tbc.alignment = right_a
    tbc.border = thin_border()
    ws1.row_dimensions[r].height = 26

    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 20

    # ═══════════════════════ Sheet 2 — Transactions ════════════════════════
    ws2 = wb.create_sheet("Transactions")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = INK
    title_banner(ws2, f"Transactions  —  {label}", f"{len(txns)} transactions this month", 7)

    headers = ["Date & Time", "Type", "Amount (฿)", "Description", "Category", "Account", "Running Total (฿)"]
    header_row = 4
    for ci, h in enumerate(headers, 1):
        cell = ws2.cell(header_row, ci, h)
        cell.font = f_header()
        cell.alignment = center
        cell.border = header_border()
    ws2.row_dimensions[header_row].height = 24

    running = 0.0
    for ri, t in enumerate(txns, header_row + 1):
        ts, typ, amt = t['timestamp'], t['type'], float(t['amount'])
        desc, cat, acc = t['description'], t['category'], t['account']
        running += amt
        is_inc, is_trans = typ == 'income', typ == 'transfer'
        amt_color = GREEN if is_inc else (INK_SOFT if is_trans else RED)
        row_bg = WHITE if ri % 2 else PANEL2
        row_cells = [
            ((ts or '')[:16].replace('T', ' '), left_a, f_cell(color=INK_SOFT), None),
            (typ.title(), center, f_cell(bold=True, color=ACCENT), None),
            (amt, right_a, f_cell(bold=True, color=amt_color), money_fmt),
            (desc or '', left_a, f_cell(color=INK), None),
            (cat or '', left_a, f_cell(color=INK_SOFT), None),
            (acc or '', left_a, f_cell(color=INK_SOFT), None),
            (running, right_a, f_cell(color=INK), money_fmt),
        ]
        for ci, (val, aln, fnt, fmt2) in enumerate(row_cells, 1):
            cell = ws2.cell(ri, ci, val)
            cell.font = fnt
            cell.alignment = aln
            cell.fill = fill(row_bg)
            cell.border = thin_border()
            if fmt2:
                cell.number_format = fmt2
        ws2.row_dimensions[ri].height = 19

    for ci, w in enumerate([19, 10, 14, 32, 16, 18, 17], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = f"A{header_row + 1}"

    # ═══════════════════════ Sheet 3 — By Category ═════════════════════════
    ws3 = wb.create_sheet("By Category")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = RED
    title_banner(ws3, f"Spending by Category  —  {label}", "Expense breakdown for the month", 3)

    cat_totals = {}
    for t in txns:
        if t['type'] == 'expense':
            cat_totals[t['category']] = cat_totals.get(t['category'], 0.0) + abs(float(t['amount']))
    sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)

    hrow = 4
    for ci, h in enumerate(["Category", "Amount (฿)", "% of Total"], 1):
        cell = ws3.cell(hrow, ci, h)
        cell.font = f_header()
        cell.alignment = center
        cell.border = header_border()
    ws3.row_dimensions[hrow].height = 24

    start_row = hrow + 1
    for idx, (cat, total) in enumerate(sorted_cats):
        ri = start_row + idx
        pct = (total / expense_total) if expense_total else 0
        row_bg = WHITE if ri % 2 else PANEL2

        cat_cell = ws3.cell(ri, 1, cat)
        cat_cell.font = f_cell(color=INK)
        cat_cell.fill = fill(row_bg)
        cat_cell.alignment = left_a
        cat_cell.border = thin_border()

        amt_cell = ws3.cell(ri, 2, total)
        amt_cell.number_format = money_fmt
        amt_cell.font = f_cell(bold=True, color=ACCENT)
        amt_cell.fill = fill(row_bg)
        amt_cell.alignment = right_a
        amt_cell.border = thin_border()

        pct_cell = ws3.cell(ri, 3, pct)
        pct_cell.number_format = pct_fmt
        pct_cell.font = f_cell(color=INK_SOFT)
        pct_cell.fill = fill(row_bg)
        pct_cell.alignment = center
        pct_cell.border = thin_border()
    end_row = start_row + len(sorted_cats) - 1

    if sorted_cats:
        ws3.conditional_formatting.add(
            f"B{start_row}:B{end_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color=ACCENT, showValue=True)
        )

    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── Report helper used by /api endpoints (kept minimal, no Telegram) ───────
def get_account_balances(user_id=None):
    """Convenience helper: list of {name, balance} dicts."""
    user_id = user_id or AUTHORIZED_USER_ID
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (user_id,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows
