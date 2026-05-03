#!/usr/bin/env python3
"""
Mike's Personal Finance Tracker Bot (v20 async)
Tracks income, expenses, balances across multiple accounts.
Supports natural language logging, photo receipts (OCR), recurring subscriptions,
weekly reports with month-over-month comparison, and Excel export.
"""

import logging
import sqlite3
import os
import re
import io
import base64
from datetime import datetime, timedelta, date
from functools import wraps

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ParseMode
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ContextTypes,
    CallbackQueryHandler, ConversationHandler,
    filters
)
from dotenv import load_dotenv
import pytz

# OCR imports
import pytesseract
from PIL import Image

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# Logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'bot.log'))
    ]
)
logger = logging.getLogger(__name__)

# Config
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
AUTHORIZED_USER_ID = int(os.getenv('AUTHORIZED_USER_ID'))
DATABASE_NAME = os.path.join('/data', 'finance.db')
BANGKOK_TZ = pytz.timezone('Asia/Bangkok')

# Optional OpenAI client for enhanced receipt processing
ai_client = None
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')
if OPENAI_API_KEY and OPENAI_API_KEY.strip():
    try:
        from openai import OpenAI
        ai_client = OpenAI()
        logger.info("OpenAI client initialized for enhanced receipt scanning.")
    except Exception as e:
        logger.warning(f"Could not initialize OpenAI client: {e}. Using OCR fallback.")

# ─── Categories ───────────────────────────────────────────────────────────────
CATEGORIES = {
    # ─── 🍜 Food & Drinks ─────────────────────────────────────────────────
    'food': ('Food & Drinks', '🍜'),
    'food & drinks': ('Food & Drinks', '🍜'),
    'restaurant': ('Food & Drinks', '🍜'),
    'street food': ('Food & Drinks', '🍜'),
    'lunch': ('Food & Drinks', '🍜'),
    'dinner': ('Food & Drinks', '🍜'),
    'breakfast': ('Food & Drinks', '🍜'),
    'brunch': ('Food & Drinks', '🍜'),
    'snack': ('Food & Drinks', '🍜'),
    'supper': ('Food & Drinks', '🍜'),
    # Fast food & chains
    'kfc': ('Food & Drinks', '🍜'),
    'bonchon': ('Food & Drinks', '🍜'),
    'mcdonalds': ('Food & Drinks', '🍜'),
    'mcdonald': ('Food & Drinks', '🍜'),
    'mcd': ('Food & Drinks', '🍜'),
    'burger king': ('Food & Drinks', '🍜'),
    'pizza hut': ('Food & Drinks', '🍜'),
    'pizza company': ('Food & Drinks', '🍜'),
    'the pizza company': ('Food & Drinks', '🍜'),
    'dominos': ('Food & Drinks', '🍜'),
    'subway': ('Food & Drinks', '🍜'),
    'mos burger': ('Food & Drinks', '🍜'),
    'texas chicken': ('Food & Drinks', '🍜'),
    'popeyes': ('Food & Drinks', '🍜'),
    'wendys': ('Food & Drinks', '🍜'),
    'five guys': ('Food & Drinks', '🍜'),
    'shake shack': ('Food & Drinks', '🍜'),
    # Thai chains & restaurants
    'mk': ('Food & Drinks', '🍜'),
    'mk suki': ('Food & Drinks', '🍜'),
    'mk restaurant': ('Food & Drinks', '🍜'),
    'sizzler': ('Food & Drinks', '🍜'),
    's&p': ('Food & Drinks', '🍜'),
    'bar b q plaza': ('Food & Drinks', '🍜'),
    'bar-b-q plaza': ('Food & Drinks', '🍜'),
    'bbq plaza': ('Food & Drinks', '🍜'),
    'shabushi': ('Food & Drinks', '🍜'),
    'fuji': ('Food & Drinks', '🍜'),
    'fuji restaurant': ('Food & Drinks', '🍜'),
    'yayoi': ('Food & Drinks', '🍜'),
    'oishi': ('Food & Drinks', '🍜'),
    'oishi ramen': ('Food & Drinks', '🍜'),
    'oishi grand': ('Food & Drinks', '🍜'),
    'coco ichibanya': ('Food & Drinks', '🍜'),
    'pepper lunch': ('Food & Drinks', '🍜'),
    'hachiban ramen': ('Food & Drinks', '🍜'),
    'gyukatsu': ('Food & Drinks', '🍜'),
    'sukishi': ('Food & Drinks', '🍜'),
    'after you': ('Food & Drinks', '🍜'),
    'swensens': ('Food & Drinks', '🍜'),
    'dairy queen': ('Food & Drinks', '🍜'),
    'dq': ('Food & Drinks', '🍜'),
    'baskin robbins': ('Food & Drinks', '🍜'),
    'cold stone': ('Food & Drinks', '🍜'),
    'krispy kreme': ('Food & Drinks', '🍜'),
    'mister donut': ('Food & Drinks', '🍜'),
    'dunkin donuts': ('Food & Drinks', '🍜'),
    'dunkin': ('Food & Drinks', '🍜'),
    'auntie annes': ('Food & Drinks', '🍜'),
    'sushi': ('Food & Drinks', '🍜'),
    'ramen': ('Food & Drinks', '🍜'),
    'som tam': ('Food & Drinks', '🍜'),
    'somtam': ('Food & Drinks', '🍜'),
    'pad thai': ('Food & Drinks', '🍜'),
    'noodle': ('Food & Drinks', '🍜'),
    'noodles': ('Food & Drinks', '🍜'),
    'rice': ('Food & Drinks', '🍜'),
    'chicken': ('Food & Drinks', '🍜'),
    'pork': ('Food & Drinks', '🍜'),
    'beef': ('Food & Drinks', '🍜'),
    'fish': ('Food & Drinks', '🍜'),
    'seafood': ('Food & Drinks', '🍜'),
    'pizza': ('Food & Drinks', '🍜'),
    'burger': ('Food & Drinks', '🍜'),
    'hotdog': ('Food & Drinks', '🍜'),
    'sandwich': ('Food & Drinks', '🍜'),
    'salad': ('Food & Drinks', '🍜'),
    'steak': ('Food & Drinks', '🍜'),
    'buffet': ('Food & Drinks', '🍜'),
    'hotpot': ('Food & Drinks', '🍜'),
    'shabu': ('Food & Drinks', '🍜'),
    'yakiniku': ('Food & Drinks', '🍜'),
    'bbq': ('Food & Drinks', '🍜'),
    'grill': ('Food & Drinks', '🍜'),
    'dim sum': ('Food & Drinks', '🍜'),
    'dimsum': ('Food & Drinks', '🍜'),
    'congee': ('Food & Drinks', '🍜'),
    'jok': ('Food & Drinks', '🍜'),
    'khao man gai': ('Food & Drinks', '🍜'),
    'boat noodle': ('Food & Drinks', '🍜'),
    'thai food': ('Food & Drinks', '🍜'),
    'japanese food': ('Food & Drinks', '🍜'),
    'korean food': ('Food & Drinks', '🍜'),
    'chinese food': ('Food & Drinks', '🍜'),
    'indian food': ('Food & Drinks', '🍜'),
    'italian food': ('Food & Drinks', '🍜'),
    'western food': ('Food & Drinks', '🍜'),
    # Drinks
    'pepsi': ('Food & Drinks', '🍜'),
    'coke': ('Food & Drinks', '🍜'),
    'coca cola': ('Food & Drinks', '🍜'),
    'sprite': ('Food & Drinks', '🍜'),
    'fanta': ('Food & Drinks', '🍜'),
    'est': ('Food & Drinks', '🍜'),
    'oishi tea': ('Food & Drinks', '🍜'),
    'ichitan': ('Food & Drinks', '🍜'),
    'bubble tea': ('Food & Drinks', '🍜'),
    'boba': ('Food & Drinks', '🍜'),
    'milk tea': ('Food & Drinks', '🍜'),
    'cha tra mue': ('Food & Drinks', '🍜'),
    'thai tea': ('Food & Drinks', '🍜'),
    'smoothie': ('Food & Drinks', '🍜'),
    'juice': ('Food & Drinks', '🍜'),
    # Delivery
    'grab food': ('Food & Drinks', '🍜'),
    'grabfood': ('Food & Drinks', '🍜'),
    'food panda': ('Food & Drinks', '🍜'),
    'foodpanda': ('Food & Drinks', '🍜'),
    'lineman': ('Food & Drinks', '🍜'),
    'line man': ('Food & Drinks', '🍜'),
    'robinhood': ('Food & Drinks', '🍜'),
    # QSR (Quick Service Restaurant) companies on receipts
    'qsr of asia': ('Food & Drinks', '🍜'),
    'qsr': ('Food & Drinks', '🍜'),
    'minor food': ('Food & Drinks', '🍜'),
    'the minor food': ('Food & Drinks', '🍜'),
    'central restaurants': ('Food & Drinks', '🍜'),
    'crg': ('Food & Drinks', '🍜'),
    # Thai company names that appear on bank receipts ("To" field)
    'yum restaurants': ('Food & Drinks', '🍜'),
    'yum': ('Food & Drinks', '🍜'),
    'mcdonald': ('Food & Drinks', '🍜'),
    'mcdonalds': ('Food & Drinks', '🍜'),
    'cpall': ('Food & Drinks', '🍜'),
    'cp all': ('Food & Drinks', '🍜'),
    'central food retail': ('Groceries', '🛒'),
    'central food': ('Groceries', '🛒'),
    'big c supercenter': ('Groceries', '🛒'),
    'siam makro': ('Groceries', '🛒'),
    'ek-chai distribution': ('Groceries', '🛒'),
    'cp axtra': ('Groceries', '🛒'),
    'true digital group': ('Subscriptions', '📱'),
    'true digital': ('Subscriptions', '📱'),
    'true corp': ('Subscriptions', '📱'),
    'true corporation': ('Subscriptions', '📱'),
    'dtac': ('Subscriptions', '📱'),
    'ais': ('Subscriptions', '📱'),
    'advanced info service': ('Subscriptions', '📱'),
    'shopee thailand': ('Shopping', '👗'),
    'shopee pay': ('Shopping', '👗'),
    'lazada express': ('Shopping', '👗'),
    'payment to shopee': ('Shopping', '👗'),
    'grab holdings': ('Transport', '🚕'),
    'grab taxi': ('Transport', '🚕'),
    'bolt technology': ('Transport', '🚕'),
    'ptt public': ('Transport', '🚕'),
    'ptt oil': ('Transport', '🚕'),
    'bts group': ('Transport', '🚕'),
    'bangkok expressway': ('Transport', '🚕'),
    'metropolitan electricity': ('Housing', '🏠'),
    'metropolitan waterworks': ('Housing', '🏠'),
    'provincial electricity': ('Housing', '🏠'),
    'provincial waterworks': ('Housing', '🏠'),
    'pea': ('Housing', '🏠'),
    'mea': ('Housing', '🏠'),
    'bumrungrad': ('Health', '💊'),
    'samitivej': ('Health', '💊'),
    'bangkok dusit medical': ('Health', '💊'),
    'bdms': ('Health', '💊'),
    'boots retail': ('Health', '💊'),
    'watsons thailand': ('Health', '💊'),
    'google asia': ('Subscriptions', '📱'),
    'google payment': ('Subscriptions', '📱'),
    'apple itunes': ('Subscriptions', '📱'),
    'spotify': ('Subscriptions', '📱'),
    'netflix': ('Subscriptions', '📱'),

    # ─── ☕ Coffee ─────────────────────────────────────────────────────────
    'coffee': ('Coffee', '☕'),
    'latte': ('Coffee', '☕'),
    'cappuccino': ('Coffee', '☕'),
    'espresso': ('Coffee', '☕'),
    'americano': ('Coffee', '☕'),
    'mocha': ('Coffee', '☕'),
    'macchiato': ('Coffee', '☕'),
    'frappe': ('Coffee', '☕'),
    'starbucks': ('Coffee', '☕'),
    'cafe amazon': ('Coffee', '☕'),
    'amazon': ('Coffee', '☕'),
    'inthanin': ('Coffee', '☕'),
    'inthanin coffee': ('Coffee', '☕'),
    'wawee': ('Coffee', '☕'),
    'wawee coffee': ('Coffee', '☕'),
    'bluecup': ('Coffee', '☕'),
    'blue cup': ('Coffee', '☕'),
    'all cafe': ('Coffee', '☕'),
    'dd coffee': ('Coffee', '☕'),
    'pacamara': ('Coffee', '☕'),
    'roots': ('Coffee', '☕'),
    'roots coffee': ('Coffee', '☕'),
    'brave roasters': ('Coffee', '☕'),
    'kaizen coffee': ('Coffee', '☕'),
    'nana coffee': ('Coffee', '☕'),
    'ceresia': ('Coffee', '☕'),
    'factory coffee': ('Coffee', '☕'),
    'graph coffee': ('Coffee', '☕'),
    'akha ama': ('Coffee', '☕'),
    'doi chaang': ('Coffee', '☕'),
    'true coffee': ('Coffee', '☕'),
    'black canyon': ('Coffee', '☕'),
    'tom n toms': ('Coffee', '☕'),
    'hollys': ('Coffee', '☕'),
    'costa coffee': ('Coffee', '☕'),
    'dean & deluca': ('Coffee', '☕'),
    'paul': ('Coffee', '☕'),
    'cafe': ('Coffee', '☕'),

    # ─── 🚕 Transport ─────────────────────────────────────────────────────
    'transport': ('Transport', '🚕'),
    'grab': ('Transport', '🚕'),
    'grab car': ('Transport', '🚕'),
    'grab bike': ('Transport', '🚕'),
    'grabcar': ('Transport', '🚕'),
    'grabbike': ('Transport', '🚕'),
    'bolt': ('Transport', '🚕'),
    'bts': ('Transport', '🚕'),
    'mrt': ('Transport', '🚕'),
    'taxi': ('Transport', '🚕'),
    'motorbike': ('Transport', '🚕'),
    'motorbike taxi': ('Transport', '🚕'),
    'bike': ('Transport', '🚕'),
    'bus': ('Transport', '🚕'),
    'van': ('Transport', '🚕'),
    'minivan': ('Transport', '🚕'),
    'songthaew': ('Transport', '🚕'),
    'tuk tuk': ('Transport', '🚕'),
    'tuktuk': ('Transport', '🚕'),
    'boat': ('Transport', '🚕'),
    'ferry': ('Transport', '🚕'),
    'chao phraya': ('Transport', '🚕'),
    'airport link': ('Transport', '🚕'),
    'airport rail': ('Transport', '🚕'),
    'srt': ('Transport', '🚕'),
    'train': ('Transport', '🚕'),
    'parking': ('Transport', '🚕'),
    'toll': ('Transport', '🚕'),
    'expressway': ('Transport', '🚕'),
    'gas': ('Transport', '🚕'),
    'petrol': ('Transport', '🚕'),
    'gasoline': ('Transport', '🚕'),
    'ptt': ('Transport', '🚕'),
    'shell': ('Transport', '🚕'),
    'bangchak': ('Transport', '🚕'),
    'caltex': ('Transport', '🚕'),
    'esso': ('Transport', '🚕'),
    'indriver': ('Transport', '🚕'),
    'cabb': ('Transport', '🚕'),
    'robinhood ride': ('Transport', '🚕'),

    # ─── 🛒 Groceries ─────────────────────────────────────────────────────
    'groceries': ('Groceries', '🛒'),
    'grocery': ('Groceries', '🛒'),
    'supermarket': ('Groceries', '🛒'),
    'market': ('Groceries', '🛒'),
    'fresh market': ('Groceries', '🛒'),
    'wet market': ('Groceries', '🛒'),
    'big c': ('Groceries', '🛒'),
    'bigc': ('Groceries', '🛒'),
    'tops': ('Groceries', '🛒'),
    'tops market': ('Groceries', '🛒'),
    'tops daily': ('Groceries', '🛒'),
    'lotus': ('Groceries', '🛒'),
    'tesco': ('Groceries', '🛒'),
    'tesco lotus': ('Groceries', '🛒'),
    'makro': ('Groceries', '🛒'),
    'villa market': ('Groceries', '🛒'),
    'villa': ('Groceries', '🛒'),
    'gourmet market': ('Groceries', '🛒'),
    'foodland': ('Groceries', '🛒'),
    'maxvalu': ('Groceries', '🛒'),
    'max value': ('Groceries', '🛒'),
    'donki': ('Groceries', '🛒'),
    'don don donki': ('Groceries', '🛒'),
    'don quijote': ('Groceries', '🛒'),
    'lawson': ('Groceries', '🛒'),
    'family mart': ('Groceries', '🛒'),
    'familymart': ('Groceries', '🛒'),
    '7-eleven': ('Groceries', '🛒'),
    '7-11': ('Groceries', '🛒'),
    '711': ('Groceries', '🛒'),
    'seven eleven': ('Groceries', '🛒'),
    'mini big c': ('Groceries', '🛒'),
    'cp fresh mart': ('Groceries', '🛒'),
    'cp freshmart': ('Groceries', '🛒'),
    'jiffy': ('Groceries', '🛒'),
    'the mall': ('Groceries', '🛒'),
    'eathai': ('Groceries', '🛒'),
    'rimping': ('Groceries', '🛒'),

    # ─── 🏠 Housing ───────────────────────────────────────────────────────
    'housing': ('Housing', '🏠'),
    'rent': ('Housing', '🏠'),
    'condo': ('Housing', '🏠'),
    'apartment': ('Housing', '🏠'),
    'utilities': ('Housing', '🏠'),
    'electric': ('Housing', '🏠'),
    'electricity': ('Housing', '🏠'),
    'electric bill': ('Housing', '🏠'),
    'pea': ('Housing', '🏠'),
    'mea': ('Housing', '🏠'),
    'water bill': ('Housing', '🏠'),
    'mwa': ('Housing', '🏠'),
    'internet': ('Housing', '🏠'),
    'wifi': ('Housing', '🏠'),
    'true internet': ('Housing', '🏠'),
    'ais fibre': ('Housing', '🏠'),
    '3bb': ('Housing', '🏠'),
    'tot': ('Housing', '🏠'),
    'nt': ('Housing', '🏠'),
    'phone bill': ('Housing', '🏠'),
    'mobile bill': ('Housing', '🏠'),
    'ais': ('Housing', '🏠'),
    'dtac': ('Housing', '🏠'),
    'true move': ('Housing', '🏠'),
    'truemove': ('Housing', '🏠'),
    'laundry': ('Housing', '🏠'),
    'cleaning': ('Housing', '🏠'),
    'maid': ('Housing', '🏠'),
    'furniture': ('Housing', '🏠'),
    'ikea': ('Housing', '🏠'),
    'home pro': ('Housing', '🏠'),
    'homepro': ('Housing', '🏠'),
    'thai watsadu': ('Housing', '🏠'),
    'baan & beyond': ('Housing', '🏠'),
    'index living': ('Housing', '🏠'),
    'sb furniture': ('Housing', '🏠'),
    'mr diy': ('Housing', '🏠'),

    # ─── 💊 Health ────────────────────────────────────────────────────────
    'health': ('Health', '💊'),
    'pharmacy': ('Health', '💊'),
    'clinic': ('Health', '💊'),
    'hospital': ('Health', '💊'),
    'doctor': ('Health', '💊'),
    'dentist': ('Health', '💊'),
    'dental': ('Health', '💊'),
    'medicine': ('Health', '💊'),
    'drug': ('Health', '💊'),
    'drugs': ('Health', '💊'),
    'vitamin': ('Health', '💊'),
    'vitamins': ('Health', '💊'),
    'multivitamin': ('Health', '💊'),
    'supplement': ('Health', '💊'),
    'supplements': ('Health', '💊'),
    # Brands
    'swisse': ('Health', '💊'),
    'dr pong': ('Health', '💊'),
    'dr.pong': ('Health', '💊'),
    'blackmores': ('Health', '💊'),
    'centrum': ('Health', '💊'),
    'vistra': ('Health', '💊'),
    'mega we care': ('Health', '💊'),
    'mega wecare': ('Health', '💊'),
    'amsel': ('Health', '💊'),
    'bioganic': ('Health', '💊'),
    'real elixir': ('Health', '💊'),
    'dhc': ('Health', '💊'),
    'nature bounty': ('Health', '💊'),
    'gnc': ('Health', '💊'),
    'now foods': ('Health', '💊'),
    'solgar': ('Health', '💊'),
    'whey protein': ('Health', '💊'),
    'protein': ('Health', '💊'),
    'collagen': ('Health', '💊'),
    'probiotic': ('Health', '💊'),
    'fish oil': ('Health', '💊'),
    'omega': ('Health', '💊'),
    'zinc': ('Health', '💊'),
    'magnesium': ('Health', '💊'),
    # Pharmacies & hospitals
    'boots': ('Health', '💊'),
    'watsons': ('Health', '💊'),
    'fascino': ('Health', '💊'),
    'bumrungrad': ('Health', '💊'),
    'samitivej': ('Health', '💊'),
    'bangkok hospital': ('Health', '💊'),
    'medpark': ('Health', '💊'),
    'praram 9': ('Health', '💊'),
    'phyathai': ('Health', '💊'),
    'paolo': ('Health', '💊'),
    'sikarin': ('Health', '💊'),
    'bnh': ('Health', '💊'),
    'eye care': ('Health', '💊'),
    'glasses': ('Health', '💊'),
    'contact lens': ('Health', '💊'),
    'sunscreen': ('Health', '💊'),
    'skincare': ('Health', '💊'),

    # ─── 👗 Shopping ──────────────────────────────────────────────────────
    'shopping': ('Shopping', '👗'),
    'clothes': ('Shopping', '👗'),
    'clothing': ('Shopping', '👗'),
    'accessories': ('Shopping', '👗'),
    'shoes': ('Shopping', '👗'),
    'sneakers': ('Shopping', '👗'),
    'bag': ('Shopping', '👗'),
    'bags': ('Shopping', '👗'),
    'watch': ('Shopping', '👗'),
    'jewelry': ('Shopping', '👗'),
    'perfume': ('Shopping', '👗'),
    'cosmetics': ('Shopping', '👗'),
    'makeup': ('Shopping', '👗'),
    # Brands & stores
    'uniqlo': ('Shopping', '👗'),
    'h&m': ('Shopping', '👗'),
    'zara': ('Shopping', '👗'),
    'muji': ('Shopping', '👗'),
    'cotton on': ('Shopping', '👗'),
    'gu': ('Shopping', '👗'),
    'nike': ('Shopping', '👗'),
    'adidas': ('Shopping', '👗'),
    'converse': ('Shopping', '👗'),
    'vans': ('Shopping', '👗'),
    'new balance': ('Shopping', '👗'),
    'charles & keith': ('Shopping', '👗'),
    'aldo': ('Shopping', '👗'),
    'sephora': ('Shopping', '👗'),
    'eveandboy': ('Shopping', '👗'),
    'eve and boy': ('Shopping', '👗'),
    'beautrium': ('Shopping', '👗'),
    'naraya': ('Shopping', '👗'),
    'jaspal': ('Shopping', '👗'),
    'cc double o': ('Shopping', '👗'),
    'gentlewoman': ('Shopping', '👗'),
    'pomelo': ('Shopping', '👗'),
    # Online shopping
    'shopee': ('Shopping', '👗'),
    'lazada': ('Shopping', '👗'),
    'shopeepay': ('Shopping', '👗'),
    'tiktok shop': ('Shopping', '👗'),
    'amazon': ('Shopping', '👗'),
    'aliexpress': ('Shopping', '👗'),
    # Tech & electronics
    'headphones': ('Shopping', '👗'),
    'earbuds': ('Shopping', '👗'),
    'airpods': ('Shopping', '👗'),
    'sennheiser': ('Shopping', '👗'),
    'sony': ('Shopping', '👗'),
    'samsung': ('Shopping', '👗'),
    'apple': ('Shopping', '👗'),
    'iphone': ('Shopping', '👗'),
    'ipad': ('Shopping', '👗'),
    'macbook': ('Shopping', '👗'),
    'laptop': ('Shopping', '👗'),
    'phone case': ('Shopping', '👗'),
    'charger': ('Shopping', '👗'),
    'power bank': ('Shopping', '👗'),
    'banana it': ('Shopping', '👗'),
    'jib': ('Shopping', '👗'),
    'power buy': ('Shopping', '👗'),
    'powerbuy': ('Shopping', '👗'),
    'it city': ('Shopping', '👗'),
    'studio 7': ('Shopping', '👗'),
    'istudio': ('Shopping', '👗'),
    # Malls
    'central': ('Shopping', '👗'),
    'centralworld': ('Shopping', '👗'),
    'siam paragon': ('Shopping', '👗'),
    'emquartier': ('Shopping', '👗'),
    'emporium': ('Shopping', '👗'),
    'iconsiam': ('Shopping', '👗'),
    'terminal 21': ('Shopping', '👗'),
    'mbk': ('Shopping', '👗'),
    'platinum': ('Shopping', '👗'),
    'pratunam': ('Shopping', '👗'),
    'chatuchak': ('Shopping', '👗'),
    'jj market': ('Shopping', '👗'),
    'robinson': ('Shopping', '👗'),
    'mega bangna': ('Shopping', '👗'),
    'fashion island': ('Shopping', '👗'),
    'future park': ('Shopping', '👗'),
    'seacon': ('Shopping', '👗'),

    # ─── 🎉 Entertainment ─────────────────────────────────────────────────
    'entertainment': ('Entertainment', '🎉'),
    'movies': ('Entertainment', '🎉'),
    'movie': ('Entertainment', '🎉'),
    'cinema': ('Entertainment', '🎉'),
    'sf cinema': ('Entertainment', '🎉'),
    'major cineplex': ('Entertainment', '🎉'),
    'major': ('Entertainment', '🎉'),
    'concert': ('Entertainment', '🎉'),
    'night out': ('Entertainment', '🎉'),
    'bar': ('Entertainment', '🎉'),
    'pub': ('Entertainment', '🎉'),
    'club': ('Entertainment', '🎉'),
    'nightclub': ('Entertainment', '🎉'),
    'karaoke': ('Entertainment', '🎉'),
    'ktv': ('Entertainment', '🎉'),
    'bowling': ('Entertainment', '🎉'),
    'arcade': ('Entertainment', '🎉'),
    'escape room': ('Entertainment', '🎉'),
    'museum': ('Entertainment', '🎉'),
    'exhibition': ('Entertainment', '🎉'),
    'art gallery': ('Entertainment', '🎉'),
    'zoo': ('Entertainment', '🎉'),
    'safari world': ('Entertainment', '🎉'),
    'siam ocean world': ('Entertainment', '🎉'),
    'madame tussauds': ('Entertainment', '🎉'),
    'kidzania': ('Entertainment', '🎉'),
    'theme park': ('Entertainment', '🎉'),
    'water park': ('Entertainment', '🎉'),
    'spa': ('Entertainment', '🎉'),
    'massage': ('Entertainment', '🎉'),
    'thai massage': ('Entertainment', '🎉'),
    'onsen': ('Entertainment', '🎉'),
    'let relax': ('Entertainment', '🎉'),
    'health land': ('Entertainment', '🎉'),
    'gym': ('Entertainment', '🎉'),
    'fitness': ('Entertainment', '🎉'),
    'fitness first': ('Entertainment', '🎉'),
    'virgin active': ('Entertainment', '🎉'),
    'jetts': ('Entertainment', '🎉'),
    'anytime fitness': ('Entertainment', '🎉'),
    'yoga': ('Entertainment', '🎉'),
    'swimming': ('Entertainment', '🎉'),
    'pool': ('Entertainment', '🎉'),
    'game': ('Entertainment', '🎉'),
    'games': ('Entertainment', '🎉'),
    'steam': ('Entertainment', '🎉'),
    'playstation': ('Entertainment', '🎉'),
    'nintendo': ('Entertainment', '🎉'),
    'beer': ('Entertainment', '🎉'),
    'wine': ('Entertainment', '🎉'),
    'whiskey': ('Entertainment', '🎉'),
    'cocktail': ('Entertainment', '🎉'),
    'alcohol': ('Entertainment', '🎉'),
    'singha': ('Entertainment', '🎉'),
    'chang': ('Entertainment', '🎉'),
    'leo': ('Entertainment', '🎉'),
    'heineken': ('Entertainment', '🎉'),

    # ─── 📱 Subscriptions ─────────────────────────────────────────────────
    'subscription': ('Subscriptions', '📱'),
    'subscriptions': ('Subscriptions', '📱'),
    'youtube': ('Subscriptions', '📱'),
    'youtube premium': ('Subscriptions', '📱'),
    'netflix': ('Subscriptions', '📱'),
    'spotify': ('Subscriptions', '📱'),
    'apple music': ('Subscriptions', '📱'),
    'disney plus': ('Subscriptions', '📱'),
    'disney+': ('Subscriptions', '📱'),
    'hbo go': ('Subscriptions', '📱'),
    'viu': ('Subscriptions', '📱'),
    'wetv': ('Subscriptions', '📱'),
    'iqiyi': ('Subscriptions', '📱'),
    'google one': ('Subscriptions', '📱'),
    'icloud': ('Subscriptions', '📱'),
    'chatgpt': ('Subscriptions', '📱'),
    'openai': ('Subscriptions', '📱'),
    'canva': ('Subscriptions', '📱'),
    'adobe': ('Subscriptions', '📱'),
    'notion': ('Subscriptions', '📱'),
    'line': ('Subscriptions', '📱'),
    'true id': ('Subscriptions', '📱'),
    'ais play': ('Subscriptions', '📱'),
    'monomax': ('Subscriptions', '📱'),

    # ─── ✈️ Travel ─────────────────────────────────────────────────────────
    'travel': ('Travel', '✈️'),
    'hotel': ('Travel', '✈️'),
    'hostel': ('Travel', '✈️'),
    'airbnb': ('Travel', '✈️'),
    'agoda': ('Travel', '✈️'),
    'booking.com': ('Travel', '✈️'),
    'flight': ('Travel', '✈️'),
    'airline': ('Travel', '✈️'),
    'airasia': ('Travel', '✈️'),
    'nok air': ('Travel', '✈️'),
    'lion air': ('Travel', '✈️'),
    'thai airways': ('Travel', '✈️'),
    'vietjet': ('Travel', '✈️'),
    'bangkok airways': ('Travel', '✈️'),
    'trip': ('Travel', '✈️'),
    'vacation': ('Travel', '✈️'),
    'holiday': ('Travel', '✈️'),
    'passport': ('Travel', '✈️'),
    'visa fee': ('Travel', '✈️'),
    'luggage': ('Travel', '✈️'),
    'suitcase': ('Travel', '✈️'),
    'klook': ('Travel', '✈️'),
    'kkday': ('Travel', '✈️'),
    'traveloka': ('Travel', '✈️'),

    # ─── 🎓 School ────────────────────────────────────────────────────────
    'school': ('School', '🎓'),
    'tuition': ('School', '🎓'),
    'books': ('School', '🎓'),
    'book': ('School', '🎓'),
    'textbook': ('School', '🎓'),
    'stationery': ('School', '🎓'),
    'language': ('School', '🎓'),
    'italian': ('School', '🎓'),
    'class': ('School', '🎓'),
    'course': ('School', '🎓'),
    'workshop': ('School', '🎓'),
    'seminar': ('School', '🎓'),
    'training': ('School', '🎓'),
    'udemy': ('School', '🎓'),
    'skillshare': ('School', '🎓'),
    'coursera': ('School', '🎓'),
    'se-ed': ('School', '🎓'),
    'b2s': ('School', '🎓'),
    'kinokuniya': ('School', '🎓'),
    'asia books': ('School', '🎓'),

    # ─── 🚬 Cigarettes ──────────────────────────────────────────────────
    'cigarette': ('Cigarettes', '🚬'),
    'cigarettes': ('Cigarettes', '🚬'),
    'cig': ('Cigarettes', '🚬'),
    'cigs': ('Cigarettes', '🚬'),
    'smoke': ('Cigarettes', '🚬'),
    'smoking': ('Cigarettes', '🚬'),
    'tobacco': ('Cigarettes', '🚬'),
    'marlboro': ('Cigarettes', '🚬'),
    'lm': ('Cigarettes', '🚬'),
    'winston': ('Cigarettes', '🚬'),
    'camel': ('Cigarettes', '🚬'),
    'lucky strike': ('Cigarettes', '🚬'),
    'dunhill': ('Cigarettes', '🚬'),
    'esse': ('Cigarettes', '🚬'),
    'krong thip': ('Cigarettes', '🚬'),
    'krongthip': ('Cigarettes', '🚬'),
    'sai fon': ('Cigarettes', '🚬'),
    'falling rain': ('Cigarettes', '🚬'),
    'wonder': ('Cigarettes', '🚬'),
    'vape': ('Cigarettes', '🚬'),
    'pod': ('Cigarettes', '🚬'),
    'iqos': ('Cigarettes', '🚬'),
    'relx': ('Cigarettes', '🚬'),

    # ─── 💵 Income ────────────────────────────────────────────────────────
    'salary': ('Salary', '💵'),
    'freelance': ('Freelance', '💼'),
    'gallery': ('Gallery Sales', '🖼️'),
    'gallery sales': ('Gallery Sales', '🖼️'),
    'artwork': ('Artwork / Commission', '🎨'),
    'sold': ('Gallery Sales', '🖼️'),
    'commission': ('Artwork / Commission', '🎨'),
    'bonus': ('Bonus', '🏆'),
    'refund': ('Cashback / Refund', '💳'),
    'cashback': ('Cashback / Refund', '💳'),
    'dividend': ('Investment', '💹'),
    'interest': ('Investment', '💹'),
    'investment': ('Investment', '💹'),
    'gift money': ('Gift Money', '🎁'),
    'angpao': ('Gift Money', '🎁'),
    'prize': ('Bonus', '🏆'),
    'business': ('Business', '🤝'),
}

CATEGORY_LIST = [
    ('🍜', 'Food & Drinks'),
    ('☕', 'Coffee'),
    ('🚕', 'Transport'),
    ('🛒', 'Groceries'),
    ('🏠', 'Housing'),
    ('💊', 'Health'),
    ('👗', 'Shopping'),
    ('🎉', 'Entertainment'),
    ('📱', 'Subscriptions'),
    ('✈️', 'Travel'),
    ('🎓', 'School'),
    ('🚬', 'Cigarettes'),
    ('🧾', 'Other'),
]

INCOME_CATEGORY_LIST = [
    ('💵', 'Salary'),
    ('💼', 'Freelance'),
    ('🖼️', 'Gallery Sales'),
    ('🎨', 'Artwork / Commission'),
    ('🏆', 'Bonus'),
    ('🎁', 'Gift Money'),
    ('💳', 'Cashback / Refund'),
    ('💹', 'Investment'),
    ('🤝', 'Business'),
    ('🧾', 'Other Income'),
]

ACCOUNT_KEYWORDS = {
    'bank': 'Bangkok Bank',
    'bangkok bank': 'Bangkok Bank',
    'bbl': 'Bangkok Bank',
    'true money': 'True Money Wallet',
    'truemoney': 'True Money Wallet',
    'true wallet': 'True Money Wallet',
    'mrt': 'MRT EMV Visa',
    'emv': 'MRT EMV Visa',
    'visa': 'MRT EMV Visa',
    'rabbit': 'Rabbit Card',
    'rabbit card': 'Rabbit Card',
    'cash': 'Cash',
    'muvmi': 'Muvmi',
    'solsot': 'Solsot Member',
    'solsot member': 'Solsot Member',
}


# ─── Database ─────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        balance REAL NOT NULL DEFAULT 0.0,
        UNIQUE(user_id, name)
    );
    CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        emoji TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        description TEXT,
        type TEXT NOT NULL,
        category TEXT NOT NULL DEFAULT 'Other',
        account TEXT NOT NULL DEFAULT 'Cash',
        timestamp DATETIME DEFAULT (datetime('now', '+7 hours'))
    );
    CREATE TABLE IF NOT EXISTS recurring_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        amount REAL NOT NULL,
        category TEXT NOT NULL DEFAULT 'Subscriptions',
        account TEXT NOT NULL DEFAULT 'Cash',
        next_due_date DATE NOT NULL,
        frequency TEXT NOT NULL DEFAULT 'monthly'
    );
    """)
    for emoji, name in CATEGORY_LIST:
        c.execute("INSERT OR IGNORE INTO categories (name, emoji) VALUES (?, ?)", (name, emoji))
    # Create accounts if they don't exist (with 0 balance — use /updatebalance to set correct values)
    # INSERT OR IGNORE means this only runs on a fresh database, never overwrites existing balances
    uid = AUTHORIZED_USER_ID
    for user_id, name in [
        (uid, 'Bangkok Bank'),
        (uid, 'True Money Wallet'),
        (uid, 'MRT EMV Visa'),
        (uid, 'Rabbit Card'),
        (uid, 'Cash'),
        (uid, 'Muvmi'),
        (uid, 'Solsot Member'),
    ]:
        c.execute("INSERT OR IGNORE INTO accounts (user_id, name, balance) VALUES (?, ?, 0.0)", (user_id, name))
    # Note: Subscriptions are managed via /addsubscription and /deletesubscription commands.
    # No hardcoded subscription seeds — the live database on Render has the real data.
    conn.commit()
    conn.close()


def get_db():
    conn = sqlite3.connect(DATABASE_NAME)
    conn.row_factory = sqlite3.Row
    return conn


# ─── Auth decorator ───────────────────────────────────────────────────────────
def restricted(func):
    @wraps(func)
    async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        user_id = update.effective_user.id if update.effective_user else None
        if user_id != AUTHORIZED_USER_ID:
            logger.warning(f"Unauthorized access attempt by {user_id}")
            # Safe reply that works for both messages and callback queries
            msg = update.effective_message
            if msg:
                await msg.reply_text("Sorry, this bot is private. 🔒")
            return
        return await func(update, context, *args, **kwargs)
    return wrapped


# ─── Helpers ──────────────────────────────────────────────────────────────────
def detect_category(text):
    text_lower = text.lower()
    sorted_keywords = sorted(CATEGORIES.keys(), key=len, reverse=True)
    for keyword in sorted_keywords:
        if keyword in text_lower:
            return CATEGORIES[keyword]
    return ('Other', '🧾')


def detect_account(text):
    text_lower = text.lower()
    sorted_keywords = sorted(ACCOUNT_KEYWORDS.keys(), key=len, reverse=True)
    for keyword in sorted_keywords:
        if keyword in text_lower:
            return ACCOUNT_KEYWORDS[keyword]
    return 'Cash'


def parse_bangkok_bank_ocr(ocr_text):
    """Parse Bangkok Bank transfer slip from OCR text."""
    result = {
        'amount': None,
        'note': None,
        'bank': None,
        'direction': 'OUT',
        'to': None,
    }

    # Detect Bangkok Bank
    if 'bangkok bank' in ocr_text.lower():
        result['bank'] = 'Bangkok Bank'

    # Extract amount - look for patterns like "6,504.00" or "204.00" near "Amount" or "THB"
    amount_patterns = [
        r'(?:Amount|amount)\s*[:\s]*([0-9,]+\.?\d*)\s*(?:THB|Baht)?',
        r'([0-9,]+\.\d{2})\s*THB',
        r'THB\s*([0-9,]+\.\d{2})',
    ]
    for pattern in amount_patterns:
        match = re.search(pattern, ocr_text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).replace(',', '')
            try:
                amount = float(amount_str)
                if amount > 0 and amount != 0.00:  # Skip fee amounts of 0.00
                    result['amount'] = amount
                    break
            except ValueError:
                continue

    # If no amount found yet, try all decimal numbers but skip 0.00 and very small amounts
    if not result['amount']:
        all_amounts = re.findall(r'([0-9]{1,3}(?:,[0-9]{3})*\.\d{2})', ocr_text)
        valid = [float(a.replace(',', '')) for a in all_amounts if float(a.replace(',', '')) > 0]
        if valid:
            result['amount'] = max(valid)

    # Extract Note field from Bangkok Bank receipts
    # OCR sometimes outputs "Note  Swisse Multivitamin" and sometimes just "Swisse Multivitamin"
    # on its own line after the Fee (0.00 THB) line.
    skip_words = ['scan', 'verify', 'reference', 'transaction', 'bank ref', 'biller',
                  'service code', 'optional', 'bank reference', 'transaction reference']

    # Method 1: Look for explicit "Note" / "Memo" / "Remark" label
    note_patterns = [
        r'Note\s*[:\s]+(.+)',
        r'Memo\s*[:\s]+(.+)',
        r'Remark\s*[:\s]+(.+)',
    ]
    for pattern in note_patterns:
        matches = re.findall(pattern, ocr_text, re.IGNORECASE)
        for note in matches:
            note = note.strip()
            if note and len(note) > 1 and not any(sw in note.lower() for sw in skip_words):
                result['note'] = note
                break
        if result['note']:
            break

    # Method 2: If no labeled Note found, look for text between Fee line and Bank reference line
    # Bangkok Bank receipt layout: ... Fee 0.00 THB ... [Note text] ... Bank reference no.
    if not result['note']:
        fee_note_pattern = r'(?:0\.00\s*THB|Fee[^\n]*0\.00)\s*\n\s*(.+?)\s*\n\s*(?:Bank reference|Transaction reference|\d{5,})'
        match = re.search(fee_note_pattern, ocr_text, re.IGNORECASE | re.DOTALL)
        if match:
            note = match.group(1).strip()
            if note and len(note) > 1 and not any(sw in note.lower() for sw in skip_words):
                result['note'] = note

    # Method 3: Try matching any known vocabulary keyword from the OCR text lines
    # This catches cases where OCR garbles the layout but the keyword is still there
    if not result['note']:
        lines = ocr_text.split('\n')
        for line in lines:
            line = line.strip()
            if not line or len(line) < 2:
                continue
            # Skip lines that are clearly not notes
            if any(sw in line.lower() for sw in skip_words + ['amount', 'thb', 'baht', 'from', 'bangkok bank', 'mr min', '171-4', 'fee']):
                continue
            if re.match(r'^[0-9,\.\s]+$', line):  # Skip pure number lines
                continue
            if re.match(r'^[0-9]{5,}', line):  # Skip reference numbers
                continue
            # Check if this line contains any known category keyword
            line_lower = line.lower()
            from collections import OrderedDict
            for keyword in sorted(CATEGORIES.keys(), key=len, reverse=True):
                if keyword in line_lower:
                    result['note'] = line
                    break
            if result['note']:
                break

    # Extract "To" field for description
    to_patterns = [
        r'To\s+([A-Z][A-Z\s\(\)]+(?:CO\.,?LTD\.?|THAILAND|COMPANY)?)',
        r'To\s+(.+?)(?:\n|Service|Biller|Reference)',
    ]
    for pattern in to_patterns:
        match = re.search(pattern, ocr_text, re.IGNORECASE)
        if match:
            to_text = match.group(1).strip()
            if to_text and len(to_text) > 1:
                result['to'] = to_text
                break

    # Detect direction - if "From" contains the user's name/account, it's OUT
    if re.search(r'From.*(?:MR\s*MIN|171-4)', ocr_text, re.IGNORECASE):
        result['direction'] = 'OUT'
    elif re.search(r'To.*(?:MR\s*MIN|171-4)', ocr_text, re.IGNORECASE):
        result['direction'] = 'IN'

    return result


# ─── Commands ─────────────────────────────────────────────────────────────────
@restricted
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📱 *Axe Finance*\n\n"
        "Hey Mike! 👋 I'm your personal finance tracker.\n\n"
        "Tap a button below or text me things like:\n"
        "• \"spent ฿150 BTS\"\n"
        "• \"received ฿5,000 from gallery\"\n\n"
        "Or send me a photo of a receipt! 📸",
        reply_markup=build_main_menu(),
        parse_mode=ParseMode.MARKDOWN
    )
    return MENU_STATE


@restricted
async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 *Commands:*\n\n"
        "/start — Welcome message\n"
        "/help — This help menu\n"
        "/balance — Show all account balances\n"
        "/report — Get financial summary\n"
        "/export — Download Excel spreadsheet\n"
        "/backup — Download the database file\n"
        "/categories — List categories\n"
        "/subscriptions — Show recurring subscriptions\n"
        "/addsubscription — Add a subscription\n"
        "/deletesubscription — Remove a subscription\n"
        "/transfer — Transfer between accounts\n"
        "/delete — Delete last transaction\n"
        "/history — Recent transactions\n"
        "/updatebalance — Update an account balance\n\n"
        "💬 *Natural language:*\n"
        "\"spent ฿150 BTS\" — logs expense\n"
        "\"received ฿5000 salary bank\" — logs income\n"
        "\"paid ฿45 coffee\" — logs expense\n\n"
        "📸 Send a receipt photo to auto-log it\n\n"
        "📊 /export — Get monthly Excel file\n\n"
        "💡 Add account name at the end: bank, truemoney, mrt, rabbit, cash",
        parse_mode=ParseMode.MARKDOWN
    )


@restricted
async def cmd_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (AUTHORIZED_USER_ID,))
    accounts = c.fetchall()
    conn.close()

    total = sum(a['balance'] for a in accounts)
    emojis = {'Bangkok Bank': '🏦', 'True Money Wallet': '📱', 'MRT EMV Visa': '🚇', 'Rabbit Card': '🐇', 'Cash': '💵'}
    lines = ["💰 *Account Balances:*\n"]
    for a in accounts:
        e = emojis.get(a['name'], '💰')
        lines.append(f"{e} {a['name']}: ฿{a['balance']:,.2f}")
    lines.append(f"\n🏧 *Total: ฿{total:,.2f}*")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


@restricted
async def cmd_categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lines = ["📂 *Expense Categories:*\n"]
    for emoji, name in CATEGORY_LIST:
        lines.append(f"{emoji} {name}")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


@restricted
async def cmd_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT * FROM transactions WHERE user_id = ? ORDER BY timestamp DESC LIMIT 10",
        (AUTHORIZED_USER_ID,)
    )
    txns = c.fetchall()
    conn.close()

    if not txns:
        await update.message.reply_text("No transactions yet.")
        return

    lines = ["📜 *Recent Transactions:*\n"]
    for t in txns:
        emoji = "💵" if t['type'] == 'income' else "💸"
        lines.append(
            f"{emoji} ฿{abs(t['amount']):,.2f} — {t['description']}\n"
            f"   📂 {t['category']} | 🏦 {t['account']} | {t['timestamp'][:16]}"
        )
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


@restricted
@restricted
async def cmd_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT id, amount, account, description, type FROM transactions WHERE user_id = ? ORDER BY id DESC LIMIT 1",
        (AUTHORIZED_USER_ID,)
    )
    txn = c.fetchone()
    if not txn:
        await update.message.reply_text("No transactions to delete.")
        conn.close()
        return

    c.execute("DELETE FROM transactions WHERE id = ? AND user_id = ?", (txn['id'], AUTHORIZED_USER_ID))
    # Reverse the balance effect correctly: subtract what was added
    c.execute(
        "UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?",
        (txn['amount'], AUTHORIZED_USER_ID, txn['account'])
    )
    conn.commit()
    conn.close()
    sign = "+" if txn['amount'] > 0 else "-"
    await update.message.reply_text(
        f"🗑 Deleted: {txn['description']}\n"
        f"💰 {sign}฿{abs(txn['amount']):,.2f} from {txn['account']}"
    )


@restricted
async def cmd_update_balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args or len(args) < 2:
        await update.message.reply_text(
            "Usage: /updatebalance <account> <amount>\n"
            "Example: /updatebalance rabbit 250"
        )
        return
    account_name = detect_account(args[0])
    try:
        new_balance = float(args[1].replace(',', ''))
    except ValueError:
        await update.message.reply_text("Invalid amount.")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "UPDATE accounts SET balance = ? WHERE user_id = ? AND name = ?",
        (new_balance, AUTHORIZED_USER_ID, account_name)
    )
    conn.commit()
    conn.close()
    await update.message.reply_text(f"✅ {account_name} balance updated to ฿{new_balance:,.2f}")


@restricted
async def cmd_subscriptions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT name, amount, account, next_due_date FROM recurring_subscriptions WHERE user_id = ? ORDER BY next_due_date",
        (AUTHORIZED_USER_ID,)
    )
    subs = c.fetchall()
    conn.close()

    if not subs:
        await update.message.reply_text("No recurring subscriptions.")
        return

    lines = ["🔄 *Recurring Subscriptions:*\n"]
    for s in subs:
        lines.append(
            f"📱 {s['name']}: ฿{s['amount']:,.2f}/month\n"
            f"  Account: {s['account']} | Next: {s['next_due_date']}"
        )
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN)


@restricted
async def cmd_add_subscription(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args or len(args) < 3:
        await update.message.reply_text(
            "Usage: /addsubscription <name> <amount> <account> [next_due_date]\n"
            "Example: /addsubscription Netflix 399 bank 2026-05-01\n"
            "Example: /addsubscription GoogleOne 30.45 mrt 2026-05-25"
        )
        return
    # Parse from the end: last arg might be date, second-to-last is account, before that is amount
    # Everything before amount is the name
    # Try to detect: if last arg looks like a date (YYYY-MM-DD), it's the due date
    next_due = None
    remaining_args = list(args)
    if len(remaining_args) >= 4 and re.match(r'^\d{4}-\d{2}-\d{2}$', remaining_args[-1]):
        next_due = remaining_args.pop()
    if len(remaining_args) < 3:
        await update.message.reply_text("Need at least: name, amount, account")
        return
    account_keyword = remaining_args.pop()
    account = detect_account(account_keyword)
    try:
        amount = float(remaining_args.pop().replace(',', ''))
    except ValueError:
        await update.message.reply_text("Invalid amount.")
        return
    name = " ".join(remaining_args)  # Everything left is the name (supports multi-word)
    if not next_due:
        next_due = str(date.today() + timedelta(days=30))

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO recurring_subscriptions (user_id, name, amount, category, account, next_due_date, frequency) "
        "VALUES (?, ?, ?, 'Subscriptions', ?, ?, 'monthly')",
        (AUTHORIZED_USER_ID, name, amount, account, next_due)
    )
    conn.commit()
    conn.close()
    await update.message.reply_text(f"✅ Subscription '{name}' added: ฿{amount:,.2f}/month from {account}, next due {next_due}")


@restricted
async def cmd_delete_subscription(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args:
        await update.message.reply_text("Usage: /deletesubscription <name>")
        return
    name = " ".join(args)
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "DELETE FROM recurring_subscriptions WHERE user_id = ? AND name LIKE ?",
        (AUTHORIZED_USER_ID, f"%{name}%")
    )
    if c.rowcount > 0:
        conn.commit()
        await update.message.reply_text(f"🗑 Subscription matching '{name}' deleted.")
    else:
        await update.message.reply_text(f"No subscription found matching '{name}'.")
    conn.close()


@restricted
async def cmd_transfer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if not args or len(args) < 3:
        await update.message.reply_text(
            "Usage: /transfer <amount> <from> <to>\n"
            "Accounts: bank, truemoney, mrt, rabbit, cash\n"
            "Example: /transfer 45 cash bank\n"
            "Example: /transfer 200 bank truemoney"
        )
        return
    try:
        amount = float(args[0].replace(',', '').replace('฿', ''))
    except ValueError:
        await update.message.reply_text("Invalid amount.")
        return
    if amount <= 0:
        await update.message.reply_text("Amount must be greater than zero.")
        return

    from_account = detect_account(" ".join(args[1:-1]) if len(args) > 3 else args[1])
    to_account = detect_account(args[-1])
    if from_account == to_account:
        await update.message.reply_text("From and To accounts can't be the same.")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, from_account))
    if not c.fetchone():
        await update.message.reply_text(f"Account '{from_account}' not found.")
        conn.close()
        return
    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, to_account))
    if not c.fetchone():
        await update.message.reply_text(f"Account '{to_account}' not found.")
        conn.close()
        return

    c.execute("UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?", (amount, AUTHORIZED_USER_ID, from_account))
    c.execute("UPDATE accounts SET balance = balance + ? WHERE user_id = ? AND name = ?", (amount, AUTHORIZED_USER_ID, to_account))
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'expense', 'Other', ?)",
        (AUTHORIZED_USER_ID, -amount, f"Transfer to {to_account}", from_account)
    )
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'income', 'Other', ?)",
        (AUTHORIZED_USER_ID, amount, f"Transfer from {from_account}", to_account)
    )
    conn.commit()
    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, from_account))
    new_from = c.fetchone()['balance']
    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, to_account))
    new_to = c.fetchone()['balance']
    conn.close()
    await update.message.reply_text(
        f"🔄 Transfer complete!\n\n"
        f"💸 {from_account}: -฿{amount:,.2f} → ฿{new_from:,.2f}\n"
        f"💵 {to_account}: +฿{amount:,.2f} → ฿{new_to:,.2f}"
    )


# ─── Export to Excel ─────────────────────────────────────────────────────────
@restricted
async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Export transactions as an Excel file. Usage: /export [month] [year]"""
    args = context.args
    now = datetime.now(BANGKOK_TZ)

    # Parse month/year from args
    if args and len(args) >= 2:
        try:
            month = int(args[0])
            year = int(args[1])
        except ValueError:
            month = now.month
            year = now.year
    elif args and len(args) == 1:
        try:
            month = int(args[0])
            year = now.year
        except ValueError:
            month = now.month
            year = now.year
    else:
        month = now.month
        year = now.year

    month_start = f"{year}-{month:02d}-01"
    if month == 12:
        month_end = f"{year + 1}-01-01"
    else:
        month_end = f"{year}-{month + 1:02d}-01"

    conn = get_db()
    c = conn.cursor()

    # Get transactions for the month
    c.execute(
        "SELECT timestamp, type, amount, description, category, account "
        "FROM transactions WHERE user_id = ? AND timestamp >= ? AND timestamp < ? "
        "ORDER BY timestamp",
        (AUTHORIZED_USER_ID, month_start, month_end)
    )
    txns = c.fetchall()

    # Get account balances
    c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (AUTHORIZED_USER_ID,))
    accounts = c.fetchall()
    conn.close()

    if not txns:
        await update.message.reply_text(f"No transactions found for {year}-{month:02d}.")
        return

    # Create Excel workbook
    wb = Workbook()

    # ── Transactions Sheet ──
    ws = wb.active
    ws.title = f"Transactions {year}-{month:02d}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    income_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    expense_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    money_format = '#,##0.00'

    # Headers
    headers = ['Date', 'Time', 'Type', 'Amount (฿)', 'Description', 'Category', 'Account']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data rows
    total_income = 0
    total_expense = 0
    for row_idx, t in enumerate(txns, 2):
        ts = t['timestamp']
        date_str = ts[:10] if ts else ''
        time_str = ts[11:16] if ts and len(ts) > 11 else ''

        ws.cell(row=row_idx, column=1, value=date_str).border = border
        ws.cell(row=row_idx, column=2, value=time_str).border = border
        ws.cell(row=row_idx, column=3, value=t['type'].capitalize()).border = border

        amount_cell = ws.cell(row=row_idx, column=4, value=abs(t['amount']))
        amount_cell.number_format = money_format
        amount_cell.border = border

        ws.cell(row=row_idx, column=5, value=t['description']).border = border
        ws.cell(row=row_idx, column=6, value=t['category']).border = border
        ws.cell(row=row_idx, column=7, value=t['account']).border = border

        # Color rows
        fill = income_fill if t['type'] == 'income' else expense_fill
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).fill = fill

        if t['type'] == 'income':
            total_income += abs(t['amount'])
        else:
            total_expense += abs(t['amount'])

    # Summary row
    summary_row = len(txns) + 3
    ws.cell(row=summary_row, column=3, value="Total Income:").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=total_income).number_format = money_format
    ws.cell(row=summary_row, column=4).font = Font(bold=True, color="28A745")

    ws.cell(row=summary_row + 1, column=3, value="Total Expenses:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=4, value=total_expense).number_format = money_format
    ws.cell(row=summary_row + 1, column=4).font = Font(bold=True, color="DC3545")

    ws.cell(row=summary_row + 2, column=3, value="Net:").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=4, value=total_income - total_expense).number_format = money_format
    ws.cell(row=summary_row + 2, column=4).font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18

    # ── Category Summary Sheet ──
    ws2 = wb.create_sheet(title="Category Summary")
    cat_headers = ['Category', 'Total Spent (฿)', 'Number of Transactions']
    for col, header in enumerate(cat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Calculate category totals
    cat_totals = {}
    for t in txns:
        if t['type'] == 'expense':
            cat = t['category']
            if cat not in cat_totals:
                cat_totals[cat] = {'total': 0, 'count': 0}
            cat_totals[cat]['total'] += abs(t['amount'])
            cat_totals[cat]['count'] += 1

    for row_idx, (cat, data) in enumerate(sorted(cat_totals.items(), key=lambda x: x[1]['total'], reverse=True), 2):
        ws2.cell(row=row_idx, column=1, value=cat).border = border
        ws2.cell(row=row_idx, column=2, value=data['total']).number_format = money_format
        ws2.cell(row=row_idx, column=2).border = border
        ws2.cell(row=row_idx, column=3, value=data['count']).border = border

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 22

    # ── Account Balances Sheet ──
    ws3 = wb.create_sheet(title="Account Balances")
    acc_headers = ['Account', 'Balance (฿)']
    for col, header in enumerate(acc_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    total_balance = 0
    for row_idx, acc in enumerate(accounts, 2):
        ws3.cell(row=row_idx, column=1, value=acc['name']).border = border
        ws3.cell(row=row_idx, column=2, value=acc['balance']).number_format = money_format
        ws3.cell(row=row_idx, column=2).border = border
        total_balance += acc['balance']

    total_row = len(accounts) + 3
    ws3.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(row=total_row, column=2, value=total_balance).number_format = money_format
    ws3.cell(row=total_row, column=2).font = Font(bold=True)

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 18

    # Save to bytes buffer
    month_name = datetime(year, month, 1).strftime('%B')
    filename = f"Finance_{month_name}_{year}.xlsx"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await update.message.reply_document(
        document=buffer,
        filename=filename,
        caption=f"📊 Financial report for {month_name} {year}\n"
                f"💵 Income: ฿{total_income:,.2f}\n"
                f"💸 Expenses: ฿{total_expense:,.2f}\n"
                f"📊 Net: ฿{total_income - total_expense:,.2f}"
    )


# ─── Natural language handler ────────────────────────────────────────────────
@restricted
@restricted
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    text_lower = text.lower().strip()

    is_expense = False
    is_income = False
    expense_words = ['spent', 'paid', 'buy', 'bought', 'pay', 'cost', 'expense']
    income_words = ['received', 'got', 'earned', 'income', 'sold', 'salary', 'freelance', 'commission', 'transfer in']

    for w in expense_words:
        if w in text_lower:
            is_expense = True
            break
    for w in income_words:
        if w in text_lower:
            is_income = True
            break

    if not is_expense and not is_income:
        await update.message.reply_text(
            "I couldn't tell if that's income or expense. 🤔\n"
            "Try: \"spent ฿150 BTS\" or \"received ฿5000 salary\""
        )
        return

    amount_match = re.search(r'[฿฿B]?\s*([\d,]+(?:\.\d{1,2})?)', text)
    if not amount_match:
        await update.message.reply_text("I couldn't find an amount. Please include something like ฿150 or 150.")
        return

    amount = float(amount_match.group(1).replace(',', ''))
    if amount <= 0:
        await update.message.reply_text("Amount must be greater than zero.")
        return

    cat_name, cat_emoji = detect_category(text)
    account_name = detect_account(text)

    desc = text
    desc = re.sub(r'[฿฿B]?\s*[\d,]+(?:\.\d{1,2})?', '', desc, count=1)
    for w in expense_words + income_words:
        desc = re.sub(r'\b' + w + r'\b', '', desc, flags=re.IGNORECASE)
    for kw in ACCOUNT_KEYWORDS:
        desc = re.sub(r'\b' + re.escape(kw) + r'\b', '', desc, flags=re.IGNORECASE)
    desc = re.sub(r'\s+', ' ', desc).strip(' .,;:-')
    if not desc:
        desc = cat_name

    txn_type = 'income' if is_income else 'expense'
    db_amount = amount if is_income else -amount

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, ?, ?, ?)",
        (AUTHORIZED_USER_ID, db_amount, desc, txn_type, cat_name, account_name)
    )
    c.execute(
        "UPDATE accounts SET balance = balance + ? WHERE user_id = ? AND name = ?",
        (db_amount, AUTHORIZED_USER_ID, account_name)
    )
    conn.commit()
    conn.close()

    if is_income:
        await update.message.reply_text(
            f"💵 Logged income: +฿{amount:,.2f}\n📝 {desc}\n🏦 {account_name}\n📂 {cat_emoji} {cat_name}"
        )
    else:
        await update.message.reply_text(
            f"💸 Logged expense: -฿{amount:,.2f}\n📝 {desc}\n🏦 {account_name}\n📂 {cat_emoji} {cat_name}"
        )


# ─── Photo receipt handling (OCR-based, free) ────────────────────────────────
@restricted
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    caption = update.message.caption or ""
    caption_lower = caption.lower().strip()

    await update.message.reply_text("📸 Processing your receipt... give me a moment.")

    try:
        photo_file = await update.message.photo[-1].get_file()
        photo_bytes = await photo_file.download_as_bytearray()

        amount = None
        transaction_description = None
        account_name = 'Cash'
        cat_name = 'Other'
        is_income = False

        # Try AI first if available
        ai_success = False
        if ai_client:
            try:
                b64_image = base64.b64encode(photo_bytes).decode('utf-8')
                response = ai_client.chat.completions.create(
                    model="gpt-4.1-mini",
                    messages=[
                        {
                            "role": "system",
                            "content": (
                                "You are a receipt/bank transfer slip parser. Extract information from the image.\n"
                                "Look for:\n"
                                "- The total amount or transfer amount\n"
                                "- The 'Note' or 'Memo' field if present\n"
                                "- The bank or source (e.g., Bangkok Bank, KBank, SCB, True Money)\n"
                                "- Whether this is money going OUT (payment/transfer/expense) or money coming IN\n\n"
                                "Respond in this exact format:\n"
                                "AMOUNT: <number only>\n"
                                "NOTE: <the Note/Memo field, or short description>\n"
                                "BANK: <bank name if visible, otherwise UNKNOWN>\n"
                                "DIRECTION: <OUT or IN>\n"
                                "CATEGORY: <one of: Food & Drinks, Coffee, Transport, Groceries, Housing, Health, Shopping, Entertainment, Subscriptions, Travel, School, Other>\n"
                                "If you cannot read the receipt, respond with AMOUNT: 0"
                            )
                        },
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": f"Extract amount, note, bank, direction from this receipt.{' Caption: ' + caption if caption else ''}"},
                                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_image}"}}
                            ]
                        }
                    ],
                    max_tokens=300
                )
                result = response.choices[0].message.content
                logger.info(f"AI receipt result: {result}")

                amount_match = re.search(r'AMOUNT:\s*([\d,]+(?:\.\d{1,2})?)', result)
                note_match = re.search(r'NOTE:\s*(.+)', result)
                bank_match = re.search(r'BANK:\s*(.+)', result)
                direction_match = re.search(r'DIRECTION:\s*(OUT|IN)', result, re.IGNORECASE)
                cat_match_ai = re.search(r'CATEGORY:\s*(.+)', result)

                if amount_match and float(amount_match.group(1).replace(',', '')) > 0:
                    amount = float(amount_match.group(1).replace(',', ''))
                    # If caption is present, it will override this later
                    transaction_description = note_match.group(1).strip() if note_match else "Receipt scan"
                    cat_name = cat_match_ai.group(1).strip() if cat_match_ai else "Other"
                    is_income = direction_match.group(1).upper() == "IN" if direction_match else False

                    if bank_match:
                        bank_text = bank_match.group(1).strip().lower()
                        if bank_text != 'unknown':
                            if 'bangkok' in bank_text or 'bbl' in bank_text:
                                account_name = 'Bangkok Bank'
                            elif 'true' in bank_text:
                                account_name = 'True Money Wallet'
                            elif 'rabbit' in bank_text:
                                account_name = 'Rabbit Card'
                            elif 'mrt' in bank_text or 'emv' in bank_text:
                                account_name = 'MRT EMV Visa'
                    ai_success = True
            except Exception as e:
                logger.warning(f"AI receipt processing failed, falling back to OCR: {e}")

        # Fallback: OCR-based parsing
        ocr_text = ''
        if not ai_success:
            try:
                image = Image.open(io.BytesIO(bytes(photo_bytes)))
                ocr_text = pytesseract.image_to_string(image, lang='eng')
                logger.info(f"OCR text: {ocr_text[:500]}")

                parsed = parse_bangkok_bank_ocr(ocr_text)

                if parsed['amount'] and parsed['amount'] > 0:
                    amount = parsed['amount']
                    # PRIORITY: Note field always wins over To field
                    # Note has the user's own description (e.g., "Swisse Multivitamin", "Kfc", "Pepsi")
                    # To field is just the recipient company name
                    if parsed["note"]:
                        transaction_description = parsed["note"]
                    elif parsed["to"]:
                        transaction_description = parsed["to"]
                    else:
                        transaction_description = "Receipt scan"
                    is_income = parsed['direction'] == 'IN'

                    if parsed['bank']:
                        if 'bangkok' in parsed['bank'].lower():
                            account_name = 'Bangkok Bank'
                else:
                    # Try harder with different amounts
                    all_amounts = re.findall(r'([0-9,]+\.\d{2})', ocr_text)
                    valid_amounts = [float(a.replace(',', '')) for a in all_amounts if float(a.replace(',', '')) > 0]
                    if valid_amounts:
                        amount = max(valid_amounts)
                        # Try to get note first
                        note_match = re.search(r'Note\s*[:\s]+([A-Za-z][A-Za-z0-9\s,\.\-\']+)', ocr_text, re.IGNORECASE)
                        if note_match:
                            note = note_match.group(1).strip()
                            skip_words = ["scan", "verify", "reference", "transaction", "bank ref"]
                            if note and not any(sw in note.lower() for sw in skip_words):
                                transaction_description = note
                            else:
                                transaction_description = "Receipt scan"
                        else:
                            transaction_description = "Receipt scan"
                        if 'bangkok bank' in ocr_text.lower():
                            account_name = 'Bangkok Bank'
            except Exception as e:
                logger.error(f"OCR processing failed: {e}")

        # If we still don't have an amount, give up
        if not amount or amount <= 0:
            await update.message.reply_text(
                "😅 I couldn't read the receipt clearly. Please log it manually:\n"
                "Example: \"spent ฿150 food bank\""
            )
            return

        # Override description and categorize based on caption if present
        if caption:
            transaction_description = caption
            detected_cat, _ = detect_category(caption)
            if detected_cat != 'Other':
                cat_name = detected_cat

        # If no caption, or caption didn't yield a category, try from OCR/AI description
        if cat_name == 'Other' and transaction_description and transaction_description != "Receipt scan":
            detected_cat, _ = detect_category(transaction_description)
            if detected_cat != 'Other':
                cat_name = detected_cat

        # If still no category, try from full OCR text
        if cat_name == 'Other' and ocr_text:
            detected_cat, _ = detect_category(ocr_text)
            if detected_cat != 'Other':
                cat_name = detected_cat

        # Caption can also override account and direction
        if caption_lower:
            detected_acc = detect_account(caption_lower)
            if detected_acc != 'Cash':
                account_name = detected_acc
            for w in ['received', 'got', 'earned', 'income', 'money in', 'salary']:
                if w in caption_lower:
                    is_income = True
                    break
            for w in ['spent', 'paid', 'money out', 'buy', 'bought']:
                if w in caption_lower:
                    is_income = False
                    break

        valid_cats = [name for _, name in CATEGORY_LIST]
        if cat_name not in valid_cats:
            cat_name = "Other"

        cat_emoji = "🧾"
        for e, n in CATEGORY_LIST:
            if n == cat_name:
                cat_emoji = e
                break

        # ─── Auto-detect internal transfers (top-ups) ─────────────────────────
        # If the receipt is FROM Bangkok Bank TO TrueMoney/Rabbit/MRT,
        # treat it as a transfer between accounts, not an expense.
        is_transfer = False
        transfer_to_account = None

        # Build a combined text to check for transfer destination
        check_text = (transaction_description or '').lower() + ' ' + ocr_text.lower()

        # Detect transfers from Bangkok Bank to other accounts
        if account_name == 'Bangkok Bank' and not is_income:
            # TrueMoney top-up detection
            if any(kw in check_text for kw in ['truemoney', 'true money', 'tmninapp', 'transfer to true money']):
                is_transfer = True
                transfer_to_account = 'True Money Wallet'
            # Rabbit Card top-up detection
            elif any(kw in check_text for kw in ['rabbit', 'rabbit card', 'rabbit line pay', 'bts top up', 'bts topup']):
                is_transfer = True
                transfer_to_account = 'Rabbit Card'
            # MRT EMV Visa top-up detection
            elif any(kw in check_text for kw in ['mrt', 'mrt card', 'mrt emv', 'mangmoom']):
                is_transfer = True
                transfer_to_account = 'MRT EMV Visa'

        if is_transfer and transfer_to_account:
            # Log as internal transfer: deduct from Bangkok Bank, add to destination
            conn = get_db()
            c = conn.cursor()
            c.execute(
                "UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?",
                (amount, AUTHORIZED_USER_ID, 'Bangkok Bank')
            )
            c.execute(
                "UPDATE accounts SET balance = balance + ? WHERE user_id = ? AND name = ?",
                (amount, AUTHORIZED_USER_ID, transfer_to_account)
            )
            c.execute(
                "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'expense', 'Other', ?)",
                (AUTHORIZED_USER_ID, -amount, f"Transfer to {transfer_to_account}", 'Bangkok Bank')
            )
            c.execute(
                "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'income', 'Other', ?)",
                (AUTHORIZED_USER_ID, amount, f"Transfer from Bangkok Bank", transfer_to_account)
            )
            conn.commit()

            # Get updated balances
            c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, 'Bangkok Bank'))
            new_bank = c.fetchone()['balance']
            c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, transfer_to_account))
            new_dest = c.fetchone()['balance']
            conn.close()

            await update.message.reply_text(
                f"🔄 Top-up detected! Logged as transfer:\n\n"
                f"💸 Bangkok Bank: -฿{amount:,.2f} → ฿{new_bank:,.2f}\n"
                f"💵 {transfer_to_account}: +฿{amount:,.2f} → ฿{new_dest:,.2f}\n\n"
                f"Wrong? Use /delete to remove it."
            )
        else:
            # Normal expense/income logging
            db_amount = amount if is_income else -amount
            txn_type = 'income' if is_income else 'expense'

            conn = get_db()
            c = conn.cursor()
            c.execute(
                "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, ?, ?, ?)",
                (AUTHORIZED_USER_ID, db_amount, transaction_description, txn_type, cat_name, account_name)
            )
            c.execute(
                "UPDATE accounts SET balance = balance + ? WHERE user_id = ? AND name = ?",
                (db_amount, AUTHORIZED_USER_ID, account_name)
            )
            conn.commit()
            conn.close()

            sign = "💵 +" if is_income else "💸 -"
            await update.message.reply_text(
                f"📸 Receipt logged!\n"
                f"{sign}฿{amount:,.2f}\n"
                f"📝 {transaction_description}\n"
                f"📂 {cat_emoji} {cat_name}\n"
                f"🏦 {account_name}\n\n"
                f"Wrong? Use /delete to remove it."
            )

    except Exception as e:
        logger.error(f"Error processing receipt: {e}")
        await update.message.reply_text(
            "😅 Something went wrong processing the receipt. Please log it manually:\n"
            "Example: \"spent ฿150 food bank\""
        )


# ─── Reports ──────────────────────────────────────────────────────────────────
def generate_report(user_id):
    conn = get_db()
    c = conn.cursor()
    now = datetime.now(BANGKOK_TZ)
    today = now.date()

    c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (user_id,))
    accounts = c.fetchall()
    total_balance = sum(a['balance'] for a in accounts)

    this_week_start = today - timedelta(days=today.weekday())
    last_week_start = this_week_start - timedelta(days=7)
    this_month_start = today.replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)

    def get_period_stats(start, end):
        c.execute(
            "SELECT COALESCE(SUM(CASE WHEN type='income' THEN amount ELSE 0 END), 0) as income, "
            "COALESCE(SUM(CASE WHEN type='expense' THEN ABS(amount) ELSE 0 END), 0) as expense "
            "FROM transactions WHERE user_id = ? AND date(timestamp) >= ? AND date(timestamp) <= ?",
            (user_id, start, end)
        )
        return c.fetchone()

    def get_category_breakdown(start, end):
        c.execute(
            "SELECT category, SUM(ABS(amount)) as total "
            "FROM transactions WHERE user_id = ? AND type = 'expense' "
            "AND date(timestamp) >= ? AND date(timestamp) <= ? "
            "GROUP BY category ORDER BY total DESC",
            (user_id, start, end)
        )
        return c.fetchall()

    this_week = get_period_stats(this_week_start, today)
    last_week = get_period_stats(last_week_start, this_week_start - timedelta(days=1))
    this_month = get_period_stats(this_month_start, today)
    last_month = get_period_stats(last_month_start, last_month_end)
    this_week_cats = get_category_breakdown(this_week_start, today)
    conn.close()

    report = f"📊 *Weekly Financial Report*\n"
    report += f"📅 {now.strftime('%A, %B %d, %Y')}\n\n"
    report += "💰 *Account Balances:*\n"
    for acc in accounts:
        emoji = {'Bangkok Bank': '🏦', 'True Money Wallet': '📱', 'MRT EMV Visa': '🚇', 'Rabbit Card': '🐇', 'Cash': '💵', 'Muvmi': '🛺', 'Solsot Member': '🎫'}.get(acc['name'], '💰')
        report += f"  {emoji} {acc['name']}: ฿{acc['balance']:,.2f}\n"
    report += f"  *Total: ฿{total_balance:,.2f}*\n\n"

    report += "📅 *This Week:*\n"
    report += f"  💵 Income: ฿{this_week['income']:,.2f}\n"
    report += f"  💸 Expenses: ฿{this_week['expense']:,.2f}\n"
    net_week = this_week['income'] - this_week['expense']
    report += f"  📊 Net: ฿{net_week:,.2f}\n\n"

    report += "🔄 *vs Last Week:*\n"
    report += f"  💵 Income: ฿{last_week['income']:,.2f}\n"
    report += f"  💸 Expenses: ฿{last_week['expense']:,.2f}\n"
    if last_week['expense'] > 0:
        pct = ((this_week['expense'] - last_week['expense']) / last_week['expense']) * 100
        d = "📈" if pct > 0 else "📉"
        report += f"  {d} Spending {abs(pct):.0f}% {'more' if pct > 0 else 'less'} than last week\n"
    report += "\n"

    report += "📅 *This Month:*\n"
    report += f"  💵 Income: ฿{this_month['income']:,.2f}\n"
    report += f"  💸 Expenses: ฿{this_month['expense']:,.2f}\n"
    net_month = this_month['income'] - this_month['expense']
    report += f"  📊 Net: ฿{net_month:,.2f}\n\n"

    report += "🔄 *vs Last Month:*\n"
    report += f"  💵 Income: ฿{last_month['income']:,.2f}\n"
    report += f"  💸 Expenses: ฿{last_month['expense']:,.2f}\n"
    if last_month['expense'] > 0:
        pct = ((this_month['expense'] - last_month['expense']) / last_month['expense']) * 100
        d = "📈" if pct > 0 else "📉"
        report += f"  {d} Spending {abs(pct):.0f}% {'more' if pct > 0 else 'less'} than last month\n"
    report += "\n"

    if this_week_cats:
        report += "📂 *This Week by Category:*\n"
        for cat in this_week_cats:
            cat_emoji = "🧾"
            for e, n in CATEGORY_LIST:
                if n == cat['category']:
                    cat_emoji = e
                    break
            report += f"  {cat_emoji} {cat['category']}: ฿{cat['total']:,.2f}\n"

    return report


@restricted
async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    report = generate_report(AUTHORIZED_USER_ID)
    await update.message.reply_text(report, parse_mode=ParseMode.MARKDOWN)


async def send_weekly_report(context: ContextTypes.DEFAULT_TYPE):
    try:
        report = generate_report(AUTHORIZED_USER_ID)
        await context.bot.send_message(
            chat_id=AUTHORIZED_USER_ID,
            text=report,
            parse_mode=ParseMode.MARKDOWN
        )
        logger.info("Weekly report sent successfully.")
    except Exception as e:
        logger.error(f"Error sending weekly report: {e}")


# ─── Recurring subscription checker ──────────────────────────────────────────
def process_due_subscriptions():
    """Check and process any overdue subscriptions.
    Returns list of (sub_name, amount, account, user_id) tuples that were processed.
    """
    conn = get_db()
    c = conn.cursor()
    today = datetime.now(BANGKOK_TZ).date()
    today_str = str(today)

    logger.info(f"Checking subscriptions for date: {today_str}")

    c.execute(
        "SELECT id, user_id, name, amount, category, account, next_due_date, frequency "
        "FROM recurring_subscriptions WHERE next_due_date <= ?",
        (today_str,)
    )
    due_subs = c.fetchall()
    processed = []

    logger.info(f"Found {len(due_subs)} due subscription(s)")

    for sub in due_subs:
        amount = sub['amount']
        logger.info(f"Processing subscription: {sub['name']} ฿{amount} from {sub['account']} (due: {sub['next_due_date']})")

        c.execute(
            "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'expense', ?, ?)",
            (sub['user_id'], -amount, f"🔄 {sub['name']} (auto)", sub['category'], sub['account'])
        )
        c.execute(
            "UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?",
            (amount, sub['user_id'], sub['account'])
        )

        # Calculate next due date
        due_date = datetime.strptime(sub['next_due_date'], "%Y-%m-%d").date()
        if sub['frequency'] == 'monthly':
            if due_date.month == 12:
                next_due = due_date.replace(year=due_date.year + 1, month=1)
            else:
                try:
                    next_due = due_date.replace(month=due_date.month + 1)
                except ValueError:
                    next_due = (due_date.replace(day=1) + timedelta(days=32)).replace(day=due_date.day)
        elif sub['frequency'] == 'weekly':
            next_due = due_date + timedelta(weeks=1)
        elif sub['frequency'] == 'yearly':
            next_due = due_date.replace(year=due_date.year + 1)
        else:
            next_due = due_date + timedelta(days=30)

        c.execute("UPDATE recurring_subscriptions SET next_due_date = ? WHERE id = ?", (str(next_due), sub['id']))
        processed.append((sub['name'], amount, sub['account'], sub['user_id']))
        logger.info(f"Subscription '{sub['name']}' logged. Next due: {next_due}")

    conn.commit()
    conn.close()
    return processed


async def check_subscriptions(context: ContextTypes.DEFAULT_TYPE):
    """Scheduled job: check and process due subscriptions, send notifications."""
    try:
        processed = process_due_subscriptions()
        for name, amount, account, user_id in processed:
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=f"🔄 Auto-logged: {name}\n💸 -฿{amount:,.2f} from {account}"
                )
            except Exception as e:
                logger.error(f"Error notifying about subscription {name}: {e}")
    except Exception as e:
        logger.error(f"Error in check_subscriptions job: {e}")


async def startup_subscription_check(app):
    """Run subscription check on bot startup to catch any missed due dates."""
    try:
        logger.info("Running startup subscription check for missed due dates...")
        processed = process_due_subscriptions()
        for name, amount, account, user_id in processed:
            try:
                await app.bot.send_message(
                    chat_id=user_id,
                    text=f"🔄 Auto-logged (missed): {name}\n💸 -฿{amount:,.2f} from {account}"
                )
            except Exception as e:
                logger.error(f"Error notifying about missed subscription {name}: {e}")
        if processed:
            logger.info(f"Startup check: processed {len(processed)} missed subscription(s)")
        else:
            logger.info("Startup check: no missed subscriptions")
    except Exception as e:
        logger.error(f"Error in startup subscription check: {e}")


# ─── Inline Keyboard Button UI ────────────────────────────────────────────────
# Conversation states for button flows
(MENU_STATE, LOG_CAT, LOG_ACCOUNT, LOG_AMOUNT,
 INCOME_CAT, INCOME_ACCOUNT, INCOME_AMOUNT,
 SETTINGS_MENU, UPD_BAL_ACCOUNT, UPD_BAL_AMOUNT,
 TRANSFER_FROM, TRANSFER_TO, TRANSFER_AMOUNT,
 SUB_MENU, SUB_ADD_NAME, SUB_ADD_AMOUNT, SUB_ADD_ACCOUNT, SUB_ADD_DATE,
 SUB_DEL_PICK, EXPORT_TYPE, EXPORT_INPUT) = range(21)

ACCOUNT_NAMES = ['Bangkok Bank', 'MRT EMV Visa', 'True Money Wallet', 'Cash', 'Rabbit Card', 'Muvmi', 'Solsot Member']


def build_main_menu():
    """Build the main menu inline keyboard."""
    keyboard = [
        [InlineKeyboardButton("💸 Log Expense", callback_data="menu_expense"),
         InlineKeyboardButton("💵 Log Income", callback_data="menu_income")],
        [InlineKeyboardButton("💰 Balance", callback_data="menu_balance"),
         InlineKeyboardButton("📜 History", callback_data="menu_history")],
        [InlineKeyboardButton("📊 Report", callback_data="menu_report"),
         InlineKeyboardButton("📋 Export", callback_data="menu_export")],
        [InlineKeyboardButton("🔄 Subscriptions", callback_data="menu_subs"),
         InlineKeyboardButton("📂 Categories", callback_data="menu_categories")],
        [InlineKeyboardButton("⚙️ Settings", callback_data="menu_settings")],
    ]
    return InlineKeyboardMarkup(keyboard)


def build_category_keyboard():
    """Build category selection keyboard."""
    cats = [
        ("🍜 Food & Drinks", "cat_Food & Drinks"),
        ("☕ Coffee", "cat_Coffee"),
        ("🚕 Transport", "cat_Transport"),
        ("🛒 Groceries", "cat_Groceries"),
        ("🏠 Housing", "cat_Housing"),
        ("💊 Health", "cat_Health"),
        ("👗 Shopping", "cat_Shopping"),
        ("🎉 Entertainment", "cat_Entertainment"),
        ("📱 Subscriptions", "cat_Subscriptions"),
        ("✈️ Travel", "cat_Travel"),
        ("🎓 School", "cat_School"),
        ("🚬 Cigarettes", "cat_Cigarettes"),
        ("🧾 Other", "cat_Other"),
    ]
    keyboard = []
    for i in range(0, len(cats), 2):
        row = [InlineKeyboardButton(cats[i][0], callback_data=cats[i][1])]
        if i + 1 < len(cats):
            row.append(InlineKeyboardButton(cats[i+1][0], callback_data=cats[i+1][1]))
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="back_main")])
    return InlineKeyboardMarkup(keyboard)


def build_income_category_keyboard():
    """Build income-specific category selection keyboard."""
    cats = [
        ("💵 Salary", "inc_cat_Salary"),
        ("💼 Freelance", "inc_cat_Freelance"),
        ("🖼️ Gallery Sales", "inc_cat_Gallery Sales"),
        ("🎨 Artwork / Commission", "inc_cat_Artwork / Commission"),
        ("🏆 Bonus", "inc_cat_Bonus"),
        ("🎁 Gift Money", "inc_cat_Gift Money"),
        ("💳 Cashback / Refund", "inc_cat_Cashback / Refund"),
        ("💹 Investment", "inc_cat_Investment"),
        ("🤝 Business", "inc_cat_Business"),
        ("🧾 Other Income", "inc_cat_Other Income"),
    ]
    keyboard = []
    for i in range(0, len(cats), 2):
        row = [InlineKeyboardButton(cats[i][0], callback_data=cats[i][1])]
        if i + 1 < len(cats):
            row.append(InlineKeyboardButton(cats[i+1][0], callback_data=cats[i+1][1]))
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="back_main")])
    return InlineKeyboardMarkup(keyboard)


def build_account_keyboard(prefix="acc"):
    """Build account selection keyboard."""
    accs = [
        ("🏦 Bangkok Bank",    f"{prefix}_Bangkok Bank"),
        ("🚇 MRT EMV Visa",    f"{prefix}_MRT EMV Visa"),
        ("📱 True Money",      f"{prefix}_True Money Wallet"),
        ("💵 Cash",            f"{prefix}_Cash"),
        ("🐇 Rabbit Card",     f"{prefix}_Rabbit Card"),
        ("🛺 Muvmi",           f"{prefix}_Muvmi"),
        ("🎫 Solsot Member",   f"{prefix}_Solsot Member"),
    ]
    keyboard = []
    for i in range(0, len(accs), 2):
        row = [InlineKeyboardButton(accs[i][0], callback_data=accs[i][1])]
        if i + 1 < len(accs):
            row.append(InlineKeyboardButton(accs[i+1][0], callback_data=accs[i+1][1]))
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="back_main")])
    return InlineKeyboardMarkup(keyboard)


def build_subs_menu():
    """Build subscriptions submenu."""
    keyboard = [
        [InlineKeyboardButton("📋 View All", callback_data="sub_view"),
         InlineKeyboardButton("➕ Add New", callback_data="sub_add")],
        [InlineKeyboardButton("🗑 Delete", callback_data="sub_delete")],
        [InlineKeyboardButton("⬅️ Back", callback_data="back_main")],
    ]
    return InlineKeyboardMarkup(keyboard)


def build_settings_menu():
    """Build settings submenu."""
    keyboard = [
        [InlineKeyboardButton("💰 Update Balance", callback_data="set_balance"),
         InlineKeyboardButton("🔄 Transfer", callback_data="set_transfer")],
        [InlineKeyboardButton("⬅️ Back", callback_data="back_main")],
    ]
    return InlineKeyboardMarkup(keyboard)


# ─── Menu command ─────────────────────────────────────────────────────────────
@restricted
async def cmd_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show the main menu with inline buttons."""
    await update.message.reply_text(
        "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
        reply_markup=build_main_menu(),
        parse_mode=ParseMode.MARKDOWN
    )
    return MENU_STATE


async def button_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle main menu button presses."""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if user_id != AUTHORIZED_USER_ID:
        return MENU_STATE

    data = query.data

    if data == "menu_expense":
        context.user_data['flow'] = 'expense'
        await query.edit_message_text(
            "💸 *Log Expense*\n\nPick a category:",
            reply_markup=build_category_keyboard(),
            parse_mode=ParseMode.MARKDOWN
        )
        return LOG_CAT

    elif data == "menu_income":
        context.user_data['flow'] = 'income'
        await query.edit_message_text(
            "💵 *Log Income*\n\nPick a category:",
            reply_markup=build_income_category_keyboard(),
            parse_mode=ParseMode.MARKDOWN
        )
        return INCOME_CAT

    elif data == "menu_balance":
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (AUTHORIZED_USER_ID,))
        accounts = c.fetchall()
        conn.close()
        total = sum(a['balance'] for a in accounts)
        emojis = {'Bangkok Bank': '🏦', 'True Money Wallet': '📱', 'MRT EMV Visa': '🚇', 'Rabbit Card': '🐇', 'Cash': '💵'}
        lines = ["💰 *Account Balances:*\n"]
        for a in accounts:
            e = emojis.get(a['name'], '💰')
            lines.append(f"{e} {a['name']}: ฿{a['balance']:,.2f}")
        lines.append(f"\n🏧 *Total: ฿{total:,.2f}*")
        keyboard = [[InlineKeyboardButton("⬅️ Back to Menu", callback_data="back_main")]]
        await query.edit_message_text(
            "\n".join(lines),
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    elif data == "menu_history":
        conn = get_db()
        c = conn.cursor()
        c.execute(
            "SELECT * FROM transactions WHERE user_id = ? ORDER BY timestamp DESC LIMIT 10",
            (AUTHORIZED_USER_ID,)
        )
        txns = c.fetchall()
        conn.close()
        if not txns:
            text = "No transactions yet."
        else:
            lines = ["📜 *Recent Transactions:*\n"]
            for t in txns:
                emoji = "💵" if t['type'] == 'income' else "💸"
                lines.append(
                    f"{emoji} ฿{abs(t['amount']):,.2f} — {t['description']}\n"
                    f"   📂 {t['category']} | 🏦 {t['account']} | {t['timestamp'][:16]}"
                )
            text = "\n".join(lines)
        keyboard = [[InlineKeyboardButton("⬅️ Back to Menu", callback_data="back_main")]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.MARKDOWN)
        return MENU_STATE

    elif data == "menu_report":
        report = generate_report(AUTHORIZED_USER_ID)
        keyboard = [[InlineKeyboardButton("⬅️ Back to Menu", callback_data="back_main")]]
        await query.edit_message_text(report, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.MARKDOWN)
        return MENU_STATE

    elif data == "menu_export":
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📅  By Month", callback_data="export_month"),
             InlineKeyboardButton("📆  By Year",  callback_data="export_year")],
            [InlineKeyboardButton("⬅️ Back", callback_data="back_main")],
        ])
        await query.edit_message_text(
            "📋 *Export to Excel*\n\nChoose what you want to export:",
            reply_markup=keyboard,
            parse_mode=ParseMode.MARKDOWN
        )
        return EXPORT_TYPE

    elif data == "menu_categories":
        lines = ["📂 *Expense Categories:*\n"]
        for emoji, name in CATEGORY_LIST:
            lines.append(f"{emoji} {name}")
        keyboard = [[InlineKeyboardButton("⬅️ Back to Menu", callback_data="back_main")]]
        await query.edit_message_text("\n".join(lines), reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.MARKDOWN)
        return MENU_STATE

    elif data == "menu_subs":
        await query.edit_message_text(
            "🔄 *Subscriptions*\n\nWhat would you like to do?",
            reply_markup=build_subs_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SUB_MENU

    elif data == "menu_settings":
        await query.edit_message_text(
            "⚙️ *Settings*\n\nChoose an option:",
            reply_markup=build_settings_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SETTINGS_MENU

    elif data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    return MENU_STATE


# ─── Log Expense flow ─────────────────────────────────────────────────────────
async def log_expense_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle category selection for expense."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    cat_name = query.data.replace("cat_", "")
    context.user_data['category'] = cat_name
    await query.edit_message_text(
        f"💸 Category: *{cat_name}*\n\nNow pick an account:",
        reply_markup=build_account_keyboard("acc"),
        parse_mode=ParseMode.MARKDOWN
    )
    return LOG_ACCOUNT


async def log_expense_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle account selection for expense."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    account_name = query.data.replace("acc_", "")
    context.user_data['account'] = account_name
    cat = context.user_data.get('category', 'Other')
    await query.edit_message_text(
        f"💸 Category: *{cat}*\n🏦 Account: *{account_name}*\n\n"
        f"Now type the amount and description:\n"
        f"Example: `150 KFC` or `45.50 coffee`",
        parse_mode=ParseMode.MARKDOWN
    )
    return LOG_AMOUNT


async def log_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle amount+description text input for expense."""
    text = update.message.text.strip()

    amount_match = re.match(r'[฿฿B]?\s*([\d,]+(?:\.\d{1,2})?)\s*(.*)', text)
    if not amount_match:
        await update.message.reply_text("Please type amount first, then description.\nExample: 150 KFC")
        return LOG_AMOUNT

    amount = float(amount_match.group(1).replace(',', ''))
    if amount <= 0:
        await update.message.reply_text("Amount must be greater than zero. Try again:")
        return LOG_AMOUNT
    desc = amount_match.group(2).strip() or context.user_data.get('category', 'Expense')

    cat_name = context.user_data.get('category', 'Other')
    account_name = context.user_data.get('account', 'Cash')

    cat_emoji = "🧾"
    for e, n in CATEGORY_LIST:
        if n == cat_name:
            cat_emoji = e
            break

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'expense', ?, ?)",
        (AUTHORIZED_USER_ID, -amount, desc, cat_name, account_name)
    )
    c.execute(
        "UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?",
        (amount, AUTHORIZED_USER_ID, account_name)
    )
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"💸 Logged expense: -฿{amount:,.2f}\n"
        f"📝 {desc}\n"
        f"📂 {cat_emoji} {cat_name}\n"
        f"🏦 {account_name}\n\n"
        f"Wrong? Use /delete to remove it.",
        reply_markup=build_main_menu()
    )
    return MENU_STATE


# ─── Log Income flow ──────────────────────────────────────────────────────────
async def log_income_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle category selection for income."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    cat_name = query.data.replace("inc_cat_", "").replace("cat_", "")
    context.user_data['category'] = cat_name
    await query.edit_message_text(
        f"💵 Category: *{cat_name}*\n\nNow pick an account:",
        reply_markup=build_account_keyboard("acc"),
        parse_mode=ParseMode.MARKDOWN
    )
    return INCOME_ACCOUNT


async def log_income_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle account selection for income."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    account_name = query.data.replace("acc_", "")
    context.user_data['account'] = account_name
    cat = context.user_data.get('category', 'Income')
    await query.edit_message_text(
        f"💵 Category: *{cat}*\n🏦 Account: *{account_name}*\n\n"
        f"Now type the amount and description:\n"
        f"Example: `5000 salary` or `1500 freelance`",
        parse_mode=ParseMode.MARKDOWN
    )
    return INCOME_AMOUNT


async def log_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle amount+description text input for income."""
    text = update.message.text.strip()

    amount_match = re.match(r'[฿฿B]?\s*([\d,]+(?:\.\d{1,2})?)\s*(.*)', text)
    if not amount_match:
        await update.message.reply_text("Please type amount first, then description.\nExample: 5000 salary")
        return INCOME_AMOUNT

    amount = float(amount_match.group(1).replace(',', ''))
    if amount <= 0:
        await update.message.reply_text("Amount must be greater than zero. Try again:")
        return INCOME_AMOUNT
    desc = amount_match.group(2).strip() or context.user_data.get('category', 'Income')

    cat_name = context.user_data.get('category', 'Income')
    account_name = context.user_data.get('account', 'Cash')

    cat_emoji = "🧾"
    for e, n in INCOME_CATEGORY_LIST + CATEGORY_LIST:
        if n == cat_name:
            cat_emoji = e
            break

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'income', ?, ?)",
        (AUTHORIZED_USER_ID, amount, desc, cat_name, account_name)
    )
    c.execute(
        "UPDATE accounts SET balance = balance + ? WHERE user_id = ? AND name = ?",
        (amount, AUTHORIZED_USER_ID, account_name)
    )
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"💵 Logged income: +฿{amount:,.2f}\n"
        f"📝 {desc}\n"
        f"📂 {cat_emoji} {cat_name}\n"
        f"🏦 {account_name}",
        reply_markup=build_main_menu()
    )
    return MENU_STATE


# ─── Subscriptions submenu ─────────────────────────────────────────────────────
async def sub_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle subscription submenu buttons."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    elif query.data == "sub_view":
        conn = get_db()
        c = conn.cursor()
        c.execute(
            "SELECT name, amount, account, next_due_date FROM recurring_subscriptions WHERE user_id = ? ORDER BY next_due_date",
            (AUTHORIZED_USER_ID,)
        )
        subs = c.fetchall()
        conn.close()
        if not subs:
            text = "No recurring subscriptions."
        else:
            lines = ["🔄 *Recurring Subscriptions:*\n"]
            for s in subs:
                lines.append(
                    f"📱 {s['name']}: ฿{s['amount']:,.2f}/month\n"
                    f"  Account: {s['account']} | Next: {s['next_due_date']}"
                )
            text = "\n".join(lines)
        keyboard = [[InlineKeyboardButton("⬅️ Back", callback_data="back_subs")]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode=ParseMode.MARKDOWN)
        return SUB_MENU

    elif query.data == "sub_add":
        await query.edit_message_text(
            "➕ *Add Subscription*\n\n"
            "Type the name of the subscription:\n"
            "Example: `Netflix` or `Spotify`",
            parse_mode=ParseMode.MARKDOWN
        )
        return SUB_ADD_NAME

    elif query.data == "sub_delete":
        conn = get_db()
        c = conn.cursor()
        c.execute(
            "SELECT id, name, amount FROM recurring_subscriptions WHERE user_id = ?",
            (AUTHORIZED_USER_ID,)
        )
        subs = c.fetchall()
        conn.close()
        if not subs:
            keyboard = [[InlineKeyboardButton("⬅️ Back", callback_data="back_subs")]]
            await query.edit_message_text("No subscriptions to delete.", reply_markup=InlineKeyboardMarkup(keyboard))
            return SUB_MENU
        keyboard = []
        for s in subs:
            keyboard.append([InlineKeyboardButton(
                f"🗑 {s['name']} (฿{s['amount']:,.2f})",
                callback_data=f"subdel_{s['id']}"
            )])
        keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="back_subs")])
        await query.edit_message_text(
            "🗑 *Delete Subscription*\n\nTap one to delete:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN
        )
        return SUB_DEL_PICK

    elif query.data == "back_subs":
        await query.edit_message_text(
            "🔄 *Subscriptions*\n\nWhat would you like to do?",
            reply_markup=build_subs_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SUB_MENU

    return SUB_MENU


async def sub_add_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get subscription name."""
    context.user_data['sub_name'] = update.message.text.strip()
    await update.message.reply_text(
        f"Name: *{context.user_data['sub_name']}*\n\n"
        f"Now type the monthly amount:\n"
        f"Example: `199` or `30.45`",
        parse_mode=ParseMode.MARKDOWN
    )
    return SUB_ADD_AMOUNT


async def sub_add_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get subscription amount."""
    try:
        amount = float(update.message.text.strip().replace(',', ''))
        context.user_data['sub_amount'] = amount
    except ValueError:
        await update.message.reply_text("Invalid amount. Please type a number like 199 or 30.45")
        return SUB_ADD_AMOUNT

    await update.message.reply_text(
        f"Name: *{context.user_data['sub_name']}*\n"
        f"Amount: *฿{amount:,.2f}*\n\n"
        f"Pick the account to charge from:",
        reply_markup=build_account_keyboard("subacc"),
        parse_mode=ParseMode.MARKDOWN
    )
    return SUB_ADD_ACCOUNT


async def sub_add_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get subscription account."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "🔄 *Subscriptions*\n\nWhat would you like to do?",
            reply_markup=build_subs_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SUB_MENU

    account_name = query.data.replace("subacc_", "")
    context.user_data['sub_account'] = account_name
    await query.edit_message_text(
        f"Name: *{context.user_data['sub_name']}*\n"
        f"Amount: *฿{context.user_data['sub_amount']:,.2f}*\n"
        f"Account: *{account_name}*\n\n"
        f"Type the next due date (YYYY-MM-DD):\n"
        f"Example: `2026-05-25`",
        parse_mode=ParseMode.MARKDOWN
    )
    return SUB_ADD_DATE


async def sub_add_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get subscription due date and save."""
    text = update.message.text.strip()
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', text):
        await update.message.reply_text("Please use format YYYY-MM-DD, like 2026-05-25")
        return SUB_ADD_DATE

    name = context.user_data['sub_name']
    amount = context.user_data['sub_amount']
    account = context.user_data['sub_account']

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO recurring_subscriptions (user_id, name, amount, category, account, next_due_date, frequency) "
        "VALUES (?, ?, ?, 'Subscriptions', ?, ?, 'monthly')",
        (AUTHORIZED_USER_ID, name, amount, account, text)
    )
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"✅ Subscription added!\n\n"
        f"📱 {name}: ฿{amount:,.2f}/month\n"
        f"🏦 {account}\n"
        f"📅 Next due: {text}",
        reply_markup=build_main_menu()
    )
    return MENU_STATE


async def sub_del_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle subscription deletion pick."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_subs":
        await query.edit_message_text(
            "🔄 *Subscriptions*\n\nWhat would you like to do?",
            reply_markup=build_subs_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SUB_MENU

    sub_id = int(query.data.replace("subdel_", ""))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name FROM recurring_subscriptions WHERE id = ?", (sub_id,))
    row = c.fetchone()
    name = row['name'] if row else "Unknown"
    c.execute("DELETE FROM recurring_subscriptions WHERE id = ? AND user_id = ?", (sub_id, AUTHORIZED_USER_ID))
    conn.commit()
    conn.close()

    await query.edit_message_text(
        f"🗑 Deleted subscription: {name}",
        reply_markup=build_main_menu()
    )
    return MENU_STATE


# ─── Settings submenu ─────────────────────────────────────────────────────────
async def settings_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle settings submenu buttons."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Axe Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    elif query.data == "set_balance":
        await query.edit_message_text(
            "💰 *Update Balance*\n\nPick the account:",
            reply_markup=build_account_keyboard("updbal"),
            parse_mode=ParseMode.MARKDOWN
        )
        return UPD_BAL_ACCOUNT

    elif query.data == "set_transfer":
        await query.edit_message_text(
            "🔄 *Transfer*\n\nPick the *FROM* account:",
            reply_markup=build_account_keyboard("trfrom"),
            parse_mode=ParseMode.MARKDOWN
        )
        return TRANSFER_FROM

    return SETTINGS_MENU


async def upd_bal_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle account selection for balance update."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "⚙️ *Settings*\n\nChoose an option:",
            reply_markup=build_settings_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SETTINGS_MENU

    account_name = query.data.replace("updbal_", "")
    context.user_data['upd_account'] = account_name

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, account_name))
    row = c.fetchone()
    conn.close()
    current = row['balance'] if row else 0

    await query.edit_message_text(
        f"💰 *Update Balance*\n\n"
        f"Account: *{account_name}*\n"
        f"Current balance: ฿{current:,.2f}\n\n"
        f"Type the new balance amount:",
        parse_mode=ParseMode.MARKDOWN
    )
    return UPD_BAL_AMOUNT


async def upd_bal_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle new balance amount."""
    try:
        new_balance = float(update.message.text.strip().replace(',', ''))
    except ValueError:
        await update.message.reply_text("Invalid amount. Please type a number.")
        return UPD_BAL_AMOUNT

    account_name = context.user_data.get('upd_account', 'Cash')
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "UPDATE accounts SET balance = ? WHERE user_id = ? AND name = ?",
        (new_balance, AUTHORIZED_USER_ID, account_name)
    )
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"✅ {account_name} balance updated to ฿{new_balance:,.2f}",
        reply_markup=build_main_menu()
    )
    return MENU_STATE


# ─── Transfer flow ────────────────────────────────────────────────────────────
async def transfer_from(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle FROM account selection for transfer."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "📱 *Mike Finance — Main Menu*\n\nTap a button below:",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    account_name = query.data.replace("trfrom_", "")
    context.user_data['transfer_from'] = account_name
    await query.edit_message_text(
        f"🔄 *Transfer*\n\n"
        f"From: *{account_name}*\n\n"
        f"Now pick the *TO* account:",
        reply_markup=build_account_keyboard("trto"),
        parse_mode=ParseMode.MARKDOWN
    )
    return TRANSFER_TO


async def transfer_to(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle TO account selection for transfer."""
    query = update.callback_query
    await query.answer()

    if query.data == "back_main":
        await query.edit_message_text(
            "⚙️ *Settings*\n\nChoose an option:",
            reply_markup=build_settings_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return SETTINGS_MENU

    account_name = query.data.replace("trto_", "")
    from_acc = context.user_data.get('transfer_from', '')

    if account_name == from_acc:
        await query.answer("Can't transfer to the same account!", show_alert=True)
        return TRANSFER_TO

    context.user_data['transfer_to'] = account_name
    await query.edit_message_text(
        f"🔄 *Transfer*\n\n"
        f"From: *{from_acc}*\n"
        f"To: *{account_name}*\n\n"
        f"Type the amount to transfer:",
        parse_mode=ParseMode.MARKDOWN
    )
    return TRANSFER_AMOUNT


async def transfer_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle transfer amount."""
    try:
        amount = float(update.message.text.strip().replace(',', ''))
    except ValueError:
        await update.message.reply_text("Invalid amount. Please type a number.")
        return TRANSFER_AMOUNT

    from_acc = context.user_data.get('transfer_from', 'Cash')
    to_acc = context.user_data.get('transfer_to', 'Cash')

    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE accounts SET balance = balance - ? WHERE user_id = ? AND name = ?", (amount, AUTHORIZED_USER_ID, from_acc))
    c.execute("UPDATE accounts SET balance = balance + ? WHERE user_id = ? AND name = ?", (amount, AUTHORIZED_USER_ID, to_acc))
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'expense', 'Other', ?)",
        (AUTHORIZED_USER_ID, -amount, f"Transfer to {to_acc}", from_acc)
    )
    c.execute(
        "INSERT INTO transactions (user_id, amount, description, type, category, account) VALUES (?, ?, ?, 'income', 'Other', ?)",
        (AUTHORIZED_USER_ID, amount, f"Transfer from {from_acc}", to_acc)
    )
    conn.commit()

    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, from_acc))
    new_from = c.fetchone()['balance']
    c.execute("SELECT balance FROM accounts WHERE user_id = ? AND name = ?", (AUTHORIZED_USER_ID, to_acc))
    new_to = c.fetchone()['balance']
    conn.close()

    await update.message.reply_text(
        f"🔄 Transfer complete!\n\n"
        f"💸 {from_acc}: -฿{amount:,.2f} → ฿{new_from:,.2f}\n"
        f"💵 {to_acc}: +฿{amount:,.2f} → ฿{new_to:,.2f}",
        reply_markup=build_main_menu()
    )
    return MENU_STATE


# ─── Fallback / Cancel ────────────────────────────────────────────────────────
async def cancel_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancel any active conversation and return to main menu."""
    if update.message:
        await update.message.reply_text(
            "Cancelled. Back to main menu.",
            reply_markup=build_main_menu()
        )
    return ConversationHandler.END


# ─── Error handler ────────────────────────────────────────────────────────────
async def error_handler(update, context: ContextTypes.DEFAULT_TYPE):
    logger.error(f"Error: {context.error}")


# ─── Export Conversation Handlers ─────────────────────────────────────────────
@restricted
async def export_type_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle export_month / export_year button press."""
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "back_main":
        await query.edit_message_text(
            "👋 *Mike Finance*\n\nWhat would you like to do?",
            reply_markup=build_main_menu(),
            parse_mode=ParseMode.MARKDOWN
        )
        return MENU_STATE

    if data == "export_month":
        context.user_data['export_mode'] = 'month'
        now = datetime.now(BANGKOK_TZ)
        example = now.strftime("%B %Y")
        await query.edit_message_text(
            f"📅 *Export by Month*\n\n"
            f"Type the month you want, e.g:\n"
            f"`{example}` · `Apr 2026` · `4 2026`\n\n"
            f"Or just type `this` for the current month.",
            parse_mode=ParseMode.MARKDOWN
        )
    else:
        context.user_data['export_mode'] = 'year'
        now = datetime.now(BANGKOK_TZ)
        await query.edit_message_text(
            f"📆 *Export by Year*\n\n"
            f"Type the year you want, e.g:\n"
            f"`{now.year}` · `{now.year - 1}`\n\n"
            f"Or just type `this` for the current year.",
            parse_mode=ParseMode.MARKDOWN
        )
    return EXPORT_INPUT


def _parse_month_input(text: str, now) -> tuple:
    """Parse natural language month input. Returns (month, year) or (None, None)."""
    import calendar as cal_mod
    text = text.strip().lower()

    if text in ('this', 'current', 'now'):
        return now.month, now.year

    month_names = {name.lower(): i for i, name in enumerate(cal_mod.month_name) if name}
    month_abbrs = {name.lower(): i for i, name in enumerate(cal_mod.month_abbr) if name}

    # Try "June 2026", "jun 2026", "june 26"
    parts = text.replace(',', ' ').split()
    if len(parts) == 2:
        a, b = parts
        # word month + year
        m = month_names.get(a) or month_abbrs.get(a)
        if m:
            try:
                yr = int(b)
                if yr < 100:
                    yr += 2000
                return m, yr
            except ValueError:
                pass
        # numeric month + year: "4 2026" or "2026 4"
        try:
            ia, ib = int(a), int(b)
            if 1 <= ia <= 12 and ib > 100:
                return ia, ib
            if 1 <= ib <= 12 and ia > 100:
                return ib, ia
        except ValueError:
            pass

    # Single word — month name only (use current year)
    if len(parts) == 1:
        m = month_names.get(parts[0]) or month_abbrs.get(parts[0])
        if m:
            return m, now.year
        try:
            m = int(parts[0])
            if 1 <= m <= 12:
                return m, now.year
        except ValueError:
            pass

    return None, None


def _parse_year_input(text: str, now) -> int:
    """Parse year input. Returns year int or None."""
    text = text.strip().lower()
    if text in ('this', 'current', 'now'):
        return now.year
    try:
        yr = int(text)
        if yr < 100:
            yr += 2000
        if 2000 <= yr <= 2100:
            return yr
    except ValueError:
        pass
    return None


@restricted
async def export_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle typed month/year for export."""
    import calendar as cal_mod
    text   = update.message.text.strip()
    mode   = context.user_data.get('export_mode', 'month')
    now    = datetime.now(BANGKOK_TZ)

    thinking = await update.message.reply_text("⏳ Generating your export...")

    if mode == 'month':
        month, year = _parse_month_input(text, now)
        if not month:
            await thinking.edit_text(
                "❌ Couldn't understand that. Try something like:\n`June 2026` · `Apr 2026` · `4 2026`",
                parse_mode=ParseMode.MARKDOWN
            )
            return EXPORT_INPUT

        month_start = f"{year}-{month:02d}-01"
        month_end   = f"{year}-{month+1:02d}-01" if month < 12 else f"{year+1}-01-01"
        label       = f"{cal_mod.month_name[month]} {year}"
        filename    = f"Finance_{cal_mod.month_name[month]}_{year}.xlsx"
        date_filter = ("timestamp >= ? AND timestamp < ?", (month_start, month_end))

    else:
        year = _parse_year_input(text, now)
        if not year:
            await thinking.edit_text(
                "❌ Couldn't understand that. Try something like:\n`2026` · `2025`",
                parse_mode=ParseMode.MARKDOWN
            )
            return EXPORT_INPUT

        label       = str(year)
        filename    = f"Finance_{year}.xlsx"
        date_filter = ("strftime('%Y', timestamp) = ?", (str(year),))

    conn = get_db()
    c    = conn.cursor()

    where_clause, where_params = date_filter
    c.execute(
        f"SELECT timestamp, type, amount, description, category, account "
        f"FROM transactions WHERE user_id = ? AND {where_clause} ORDER BY timestamp",
        (AUTHORIZED_USER_ID, *where_params)
    )
    txns = c.fetchall()
    c.execute("SELECT name, balance FROM accounts WHERE user_id = ? ORDER BY id", (AUTHORIZED_USER_ID,))
    accounts = c.fetchall()
    conn.close()

    if not txns:
        await thinking.edit_text(f"📭 No transactions found for *{label}*.", parse_mode=ParseMode.MARKDOWN)
        return EXPORT_INPUT

    # ── Build Excel ──────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    import io, calendar as cal_mod

    wb = Workbook()

    # ── Styles ──
    DARK      = "0D1117"
    GOLD      = "F0B429"
    GREEN     = "1FD6A0"
    RED       = "F0645A"
    MID       = "8896A8"
    LIGHT     = "DDE3ED"
    SURFACE   = "0F1320"

    def hdr_font(size=11, bold=True, color=LIGHT):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def cell_font(size=10, bold=False, color=LIGHT):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border(color="1E2530"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    money_fmt = '#,##0.00'
    center    = Alignment(horizontal="center", vertical="center")
    right_al  = Alignment(horizontal="right",  vertical="center")
    left_al   = Alignment(horizontal="left",   vertical="center")

    # ── Sheet 1: Transactions ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transactions"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value         = f"Mike Finance — {label}"
    title_cell.font          = Font(bold=True, color=GOLD, size=16, name="Calibri")
    title_cell.fill          = fill(DARK)
    title_cell.alignment     = center
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value     = f"Generated {now.strftime('%d %b %Y %H:%M')} (Bangkok Time)"
    sub.font      = Font(color=MID, size=9, name="Calibri")
    sub.fill      = fill(DARK)
    sub.alignment = center
    ws.row_dimensions[2].height = 18

    # Blank spacer
    for col in range(1, 8):
        ws.cell(3, col).fill = fill(DARK)
    ws.row_dimensions[3].height = 6

    # Headers
    headers = ["Date & Time", "Type", "Amount (฿)", "Description", "Category", "Account", "Running Balance"]
    header_row = 4
    for ci, h in enumerate(headers, 1):
        cell          = ws.cell(header_row, ci, h)
        cell.font     = hdr_font(size=10, color="000000")
        cell.fill     = fill(GOLD)
        cell.alignment = center
        cell.border   = border(GOLD)
    ws.row_dimensions[header_row].height = 22

    # Data rows
    running = 0.0
    income_total  = 0.0
    expense_total = 0.0

    for ri, txn in enumerate(txns, header_row + 1):
        ts, typ, amt, desc, cat, acc = txn
        is_inc = typ == 'income'
        val    = abs(float(amt))
        running += val if is_inc else -val
        if is_inc: income_total  += val
        else:      expense_total += val

        row_fill   = fill("131B27") if ri % 2 == 0 else fill(SURFACE)
        amt_color  = GREEN if is_inc else RED
        row_data   = [
            (ts[:16].replace("T", " ") if ts else "", left_al,   cell_font(color=MID),          None),
            (typ.title(),                              center,    cell_font(color=GOLD, bold=True), None),
            (val if is_inc else -val,                  right_al,  cell_font(color=amt_color, bold=True), money_fmt),
            (desc or "",                               left_al,   cell_font(color=LIGHT),         None),
            (cat or "",                                left_al,   cell_font(color=MID),           None),
            (acc or "",                                left_al,   cell_font(color=MID),           None),
            (running,                                  right_al,  cell_font(color=LIGHT),         money_fmt),
        ]
        for ci, (val2, aln, fnt, fmt2) in enumerate(row_data, 1):
            cell           = ws.cell(ri, ci, val2)
            cell.fill      = row_fill
            cell.font      = fnt
            cell.alignment = aln
            cell.border    = border()
            if fmt2:
                cell.number_format = fmt2
        ws.row_dimensions[ri].height = 18

    # Column widths
    col_widths = [20, 10, 14, 32, 18, 20, 18]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze panes
    ws.freeze_panes = "A5"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = GREEN

    def summary_block(ws, row, label, value, color):
        ws.merge_cells(f"A{row}:B{row}")
        lc = ws.cell(row, 1, label)
        lc.font      = Font(bold=True, color=MID, size=10, name="Calibri")
        lc.fill      = fill(SURFACE)
        lc.alignment = left_al
        lc.border    = border()
        vc = ws.cell(row, 3, value)
        vc.font           = Font(bold=True, color=color, size=12, name="Calibri")
        vc.fill           = fill(SURFACE)
        vc.alignment      = right_al
        vc.border         = border()
        vc.number_format  = money_fmt
        ws.row_dimensions[row].height = 24

    # Title
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value     = f"Summary — {label}"
    t2.font      = Font(bold=True, color=GOLD, size=14, name="Calibri")
    t2.fill      = fill(DARK)
    t2.alignment = center
    ws2.row_dimensions[1].height = 30

    for col in range(1, 4):
        ws2.cell(2, col).fill = fill(DARK)
    ws2.row_dimensions[2].height = 8

    summary_block(ws2, 3,  "Total Income",   income_total,              GREEN)
    summary_block(ws2, 4,  "Total Expenses",  expense_total,             RED)
    summary_block(ws2, 5,  "Net Savings",     income_total - expense_total,
                  GREEN if income_total >= expense_total else RED)

    for col in range(1, 4):
        ws2.cell(6, col).fill = fill(DARK)
    ws2.row_dimensions[6].height = 8

    # Account balances
    ab_hdr = ws2.cell(7, 1, "Account Balances")
    ab_hdr.font      = Font(bold=True, color=GOLD, size=10, name="Calibri")
    ab_hdr.fill      = fill(fill("1A2235").fgColor)
    ws2.merge_cells("A7:C7")
    ab_hdr.alignment = center
    ws2.row_dimensions[7].height = 20

    for ri2, (aname, abal) in enumerate(accounts, 8):
        ws2.merge_cells(f"A{ri2}:B{ri2}")
        nc = ws2.cell(ri2, 1, aname)
        nc.font = cell_font(color=LIGHT); nc.fill = fill(SURFACE)
        nc.alignment = left_al; nc.border = border()
        bc = ws2.cell(ri2, 3, float(abal))
        bc.font = cell_font(color=GOLD, bold=True); bc.fill = fill(SURFACE)
        bc.alignment = right_al; bc.border = border()
        bc.number_format = money_fmt
        ws2.row_dimensions[ri2].height = 20

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16

    # ── Sheet 3: Category Breakdown ───────────────────────────────────────
    ws3 = wb.create_sheet("By Category")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = RED

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = f"Spending by Category — {label}"
    t3.font  = Font(bold=True, color=GOLD, size=14, name="Calibri")
    t3.fill  = fill(DARK); t3.alignment = center
    ws3.row_dimensions[1].height = 30

    # Build category totals from txns
    cat_totals = {}
    for ts, typ, amt, desc, cat, acc in txns:
        if typ == 'expense':
            cat_totals[cat] = cat_totals.get(cat, 0) + abs(float(amt))
    sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)

    hrow = 2
    for ci2, h in enumerate(["Category", "Amount (฿)", "% of Total"], 1):
        cell          = ws3.cell(hrow, ci2, h)
        cell.font     = hdr_font(size=10, color="000000")
        cell.fill     = fill(GOLD)
        cell.alignment = center
        cell.border   = border(GOLD)
    ws3.row_dimensions[hrow].height = 22

    for ri3, (cat, total) in enumerate(sorted_cats, hrow + 1):
        pct = (total / expense_total * 100) if expense_total else 0
        row_fill3 = fill("131B27") if ri3 % 2 == 0 else fill(SURFACE)
        for ci3, (v3, fmt3, aln3) in enumerate([
            (cat, None, left_al),
            (total, money_fmt, right_al),
            (pct / 100, "0.0%", center),
        ], 1):
            cell3 = ws3.cell(ri3, ci3, v3)
            cell3.font = cell_font(color=LIGHT); cell3.fill = row_fill3
            cell3.alignment = aln3; cell3.border = border()
            if fmt3: cell3.number_format = fmt3
        ws3.row_dimensions[ri3].height = 18

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 14

    # ── Save & send ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    net    = income_total - expense_total
    net_em = "✅" if net >= 0 else "⚠️"
    caption = (
        f"📊 *{label} Export*\n\n"
        f"💚 Income:   ฿{income_total:,.2f}\n"
        f"🔴 Expenses: ฿{expense_total:,.2f}\n"
        f"{net_em} Net:      ฿{net:+,.2f}\n\n"
        f"_{len(txns)} transactions · 3 sheets_"
    )

    await thinking.delete()
    await update.message.reply_document(
        document=buf,
        filename=filename,
        caption=caption,
        parse_mode=ParseMode.MARKDOWN
    )

    # Offer another export
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Export another month", callback_data="export_month"),
         InlineKeyboardButton("📆 Export a year",       callback_data="export_year")],
        [InlineKeyboardButton("🏠 Back to Menu", callback_data="back_main")],
    ])
    await update.message.reply_text(
        "Export done! Want another?",
        reply_markup=keyboard
    )
    return EXPORT_TYPE


# ─── Backup command ───────────────────────────────────────────────────────────
@restricted
async def cmd_backup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send the live database file to the user via Telegram."""
    try:
        db_path = DATABASE_NAME
        if not os.path.exists(db_path):
            await update.message.reply_text("⚠️ Database file not found.")
            return
        size_kb = os.path.getsize(db_path) / 1024
        now_str = datetime.now(BANGKOK_TZ).strftime('%Y%m%d_%H%M')
        filename = f"finance_backup_{now_str}.db"
        with open(db_path, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=filename,
                caption=(
                    f"💾 *Database Backup*\n"
                    f"📅 {datetime.now(BANGKOK_TZ).strftime('%d %b %Y %H:%M')} (Bangkok)\n"
                    f"📦 Size: {size_kb:.1f} KB\n\n"
                    f"Keep this safe — it contains all your transactions.\n"
                    f"Upload it to GitHub to replace finance.db if needed."
                ),
                parse_mode=ParseMode.MARKDOWN
            )
    except Exception as e:
        logger.error(f"Backup error: {e}")
        await update.message.reply_text(f"❌ Backup failed: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    init_db()
    logger.info("Database initialized.")

    async def post_init(application):
        """Called after the application is initialized but before polling starts."""
        await startup_subscription_check(application)

    app = (
        Application.builder()
        .token(TELEGRAM_BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .write_timeout(30)
        .pool_timeout(30)
        .post_init(post_init)
        .build()
    )

    # Error handler
    app.add_error_handler(error_handler)

    # Conversation handler for button-based UI (must be added BEFORE plain text handler)
    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", cmd_start),
            CommandHandler("menu", cmd_menu),
            CallbackQueryHandler(button_main_menu, pattern=r'^menu_|^back_main$'),
        ],
        states={
            MENU_STATE: [
                CallbackQueryHandler(button_main_menu, pattern=r'^menu_|^back_main$'),
            ],
            LOG_CAT: [
                CallbackQueryHandler(log_expense_category, pattern=r'^cat_|^back_main$'),
            ],
            LOG_ACCOUNT: [
                CallbackQueryHandler(log_expense_account, pattern=r'^acc_|^back_main$'),
            ],
            LOG_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, log_expense_amount),
            ],
            INCOME_CAT: [
                CallbackQueryHandler(log_income_category, pattern=r'^inc_cat_|^back_main$'),
            ],
            INCOME_ACCOUNT: [
                CallbackQueryHandler(log_income_account, pattern=r'^acc_|^back_main$'),
            ],
            INCOME_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, log_income_amount),
            ],
            SETTINGS_MENU: [
                CallbackQueryHandler(settings_menu_handler, pattern=r'^set_|^back_main$'),
            ],
            UPD_BAL_ACCOUNT: [
                CallbackQueryHandler(upd_bal_account, pattern=r'^updbal_|^back_main$'),
            ],
            UPD_BAL_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, upd_bal_amount),
            ],
            TRANSFER_FROM: [
                CallbackQueryHandler(transfer_from, pattern=r'^trfrom_|^back_main$'),
            ],
            TRANSFER_TO: [
                CallbackQueryHandler(transfer_to, pattern=r'^trto_|^back_main$'),
            ],
            TRANSFER_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_amount),
            ],
            SUB_MENU: [
                CallbackQueryHandler(sub_menu_handler, pattern=r'^sub_|^back_main$|^back_subs$'),
            ],
            SUB_ADD_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, sub_add_name),
            ],
            SUB_ADD_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, sub_add_amount),
            ],
            SUB_ADD_ACCOUNT: [
                CallbackQueryHandler(sub_add_account, pattern=r'^subacc_|^back_main$'),
            ],
            SUB_ADD_DATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, sub_add_date),
            ],
            SUB_DEL_PICK: [
                CallbackQueryHandler(sub_del_pick, pattern=r'^subdel_|^back_subs$'),
            ],
            EXPORT_TYPE: [
                CallbackQueryHandler(export_type_handler, pattern=r'^export_month$|^export_year$|^back_main$'),
            ],
            EXPORT_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, export_input_handler),
                CallbackQueryHandler(export_type_handler, pattern=r'^export_month$|^export_year$|^back_main$'),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel_conversation),
            CommandHandler("menu", cmd_menu),
            CommandHandler("start", cmd_start),
        ],
        per_message=False,
    )
    app.add_handler(conv_handler)

    # Command handlers (these work independently of the conversation)
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("balance", cmd_balance))
    app.add_handler(CommandHandler("report", cmd_report))
    app.add_handler(CommandHandler("export", cmd_export))
    app.add_handler(CommandHandler("categories", cmd_categories))
    app.add_handler(CommandHandler("history", cmd_history))
    app.add_handler(CommandHandler("delete", cmd_delete))
    app.add_handler(CommandHandler("updatebalance", cmd_update_balance))
    app.add_handler(CommandHandler("subscriptions", cmd_subscriptions))
    app.add_handler(CommandHandler("addsubscription", cmd_add_subscription))
    app.add_handler(CommandHandler("deletesubscription", cmd_delete_subscription))
    app.add_handler(CommandHandler("transfer", cmd_transfer))
    app.add_handler(CommandHandler("backup", cmd_backup))

    # Photo handler
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))

    # Text handler (must be last — catches natural language when NOT in a conversation)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    # Scheduled jobs using v20 JobQueue
    job_queue = app.job_queue

    # Daily subscription check at 8 AM Bangkok time
    target_time_subs = datetime.now(BANGKOK_TZ).replace(hour=8, minute=0, second=0, microsecond=0).timetz()
    job_queue.run_daily(check_subscriptions, time=target_time_subs, name='check_subs')

    # Note: startup subscription check is handled by post_init via startup_subscription_check()
    # which runs before polling starts. A separate run_once would double-log subscriptions.

    # Weekly report every Monday at 9 AM Bangkok time
    target_time_report = datetime.now(BANGKOK_TZ).replace(hour=9, minute=0, second=0, microsecond=0).timetz()
    job_queue.run_daily(send_weekly_report, time=target_time_report, days=(1,), name='weekly_report')  # 1 = Monday (0=Sunday in PTB v20)

    logger.info("Scheduler configured. Weekly reports: Monday 9AM, Sub checks: daily 8AM (Bangkok time)")

    # Run with polling
    logger.info("Starting bot polling...")
    app.run_polling(
        drop_pending_updates=True,
        allowed_updates=Update.ALL_TYPES,
        poll_interval=1.0,
        timeout=30,
    )


if __name__ == '__main__':
    main()
